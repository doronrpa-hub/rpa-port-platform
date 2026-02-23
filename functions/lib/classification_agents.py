# Session 15: Multi-model optimization (Claude + Gemini)
"""
RCB Multi-Agent Classification System
Queries Firestore for Israeli tariff data, ministry requirements, and rules
Outputs: HTML email + Excel

UPDATED: Session 10 - Integrated with Module 4, 5, 6
UPDATED: Session 11 - HS code validation against tariff database
UPDATED: Session 11 - Standardized English subject line format
UPDATED: Session 15 - Multi-model optimization:
  - Agent 2 (classification): Claude Sonnet 4.5 (best reasoning)
  - Agent 6 (synthesis): Gemini 2.5 Pro (good Hebrew, lower cost)
  - Agents 1,3,4,5: Gemini 2.5 Flash (simple tasks, ~95% cheaper)
"""
import json
import re
import time
import requests
import base64
import io
import random
import string
from datetime import datetime
from lib.librarian import (
    search_all_knowledge, 
    search_extended_knowledge, 
    full_knowledge_search, 
    build_classification_context, 
    get_israeli_hs_format,
    validate_and_correct_classifications,  # Session 11: HS validation
)

# Session 10: Import new modules
try:
    from lib.invoice_validator import validate_invoice, quick_validate, FIELD_DEFINITIONS
    from lib.clarification_generator import (
        generate_missing_docs_request,
        generate_origin_request,
        DocumentType,
        UrgencyLevel,
    )
    from lib.rcb_orchestrator import process_and_respond
    MODULES_AVAILABLE = True
except ImportError as e:
    print(f"âš ï¸ New modules not available: {e}")
    MODULES_AVAILABLE = False


# Session 14: Language tools integration
try:
    from lib.language_tools import (
        HebrewLanguageChecker,
        build_rcb_subject as build_rcb_subject_v2,
        process_outgoing_text,
        create_language_toolkit,
    )
    _lang_checker = HebrewLanguageChecker()
    LANGUAGE_TOOLS_AVAILABLE = True
except ImportError as e:
    print(f"âš ï¸ Language tools not available: {e}")
    LANGUAGE_TOOLS_AVAILABLE = False

# Session 17 Phase 2: Enrichment agent integration
try:
    from lib.enrichment_agent import create_enrichment_agent
    ENRICHMENT_AVAILABLE = True
except ImportError as e:
    print(f"Enrichment agent not available: {e}")
    ENRICHMENT_AVAILABLE = False

# Intelligence module: system's own brain (Firestore-only, no AI)
try:
    from lib.intelligence import pre_classify, lookup_regulatory, lookup_fta, validate_documents, query_free_import_order, route_to_ministries
    INTELLIGENCE_AVAILABLE = True
except ImportError as e:
    print(f"Intelligence module not available: {e}")
    INTELLIGENCE_AVAILABLE = False

# Document parser: per-document type identification & field extraction
try:
    from lib.document_parser import parse_all_documents
    DOCUMENT_PARSER_AVAILABLE = True
except ImportError as e:
    print(f"Document parser not available: {e}")
    DOCUMENT_PARSER_AVAILABLE = False

# Smart question engine: elimination-based clarification
try:
    from lib.smart_questions import should_ask_questions, generate_smart_questions, format_questions_html
    SMART_QUESTIONS_AVAILABLE = True
except ImportError as e:
    print(f"Smart questions not available: {e}")
    SMART_QUESTIONS_AVAILABLE = False

# Verification loop: verify, enrich, cache every classification
try:
    from lib.verification_loop import verify_all_classifications, learn_from_verification
    VERIFICATION_AVAILABLE = True
except ImportError as e:
    print(f"Verification loop not available: {e}")
    VERIFICATION_AVAILABLE = False

# Document tracker: shipment phase detection + document completeness
try:
    from lib.document_tracker import create_tracker, Document, DocumentType as TrackerDocType, feed_parsed_documents
    TRACKER_AVAILABLE = True
except ImportError as e:
    print(f"Document tracker not available: {e}")
    TRACKER_AVAILABLE = False

# Tool-calling classification engine (Session 22)
try:
    from lib.tool_calling_engine import tool_calling_classify
    TOOL_CALLING_AVAILABLE = True
except ImportError as e:
    print(f"Tool-calling engine not available: {e}")
    TOOL_CALLING_AVAILABLE = False

# Three-way AI cross-check (Session 27)
try:
    from lib.cross_checker import cross_check_all_items, get_cross_check_summary
    CROSS_CHECK_AVAILABLE = True
except ImportError as e:
    print(f"Cross-checker not available: {e}")
    CROSS_CHECK_AVAILABLE = False

# Elimination engine: deterministic tariff tree walker (Session 33, Block D)
try:
    from lib.elimination_engine import eliminate, candidates_from_pre_classify, make_product_info
    ELIMINATION_AVAILABLE = True
except ImportError as e:
    print(f"Elimination engine not available: {e}")
    ELIMINATION_AVAILABLE = False

# Verification engine: Block E phases 4+5+flagging
try:
    from lib.verification_engine import run_verification_engine, build_verification_flags_html
    VERIFICATION_ENGINE_AVAILABLE = True
except ImportError as e:
    print(f"Verification engine not available: {e}")
    VERIFICATION_ENGINE_AVAILABLE = False

# Session 51: Embedded customs law expertise â€” broker's brain before any search
try:
    from lib.customs_law import format_legal_context_for_prompt, KNOWN_FAILURES
    CUSTOMS_LAW_AVAILABLE = True
except ImportError as e:
    print(f"Customs law module not available: {e}")
    CUSTOMS_LAW_AVAILABLE = False
    KNOWN_FAILURES = []

# =============================================================================
# FEATURE FLAGS (Session 26/27: cost optimization + cross-check)
# =============================================================================

PRE_CLASSIFY_BYPASS_ENABLED = True    # Skip AI pipeline when pre_classify confidence >= 90
PRE_CLASSIFY_BYPASS_THRESHOLD = 90    # Minimum confidence to bypass
COST_TRACKING_ENABLED = True          # Print per-call cost estimates
CROSS_CHECK_ENABLED = True            # Session 27: Run 3-way cross-check after classification
ELIMINATION_ENABLED = True            # Session 33 D9: Run elimination engine between pre_classify and Agent 2
VERIFICATION_ENGINE_ENABLED = True    # Session 34 Block E: Phase 4+5+Flagging

# Session 48: Gemini quota fast-fail â€” skip all Gemini calls after first 429
# CRIT-2 fix: timestamp instead of bare boolean â€” auto-resets after 60s
_gemini_quota_exhausted_at = 0  # epoch timestamp; 0 = not exhausted

# Session 40: Shared design constants from tracker email
try:
    from lib.tracker_email import (
        _RPA_BLUE, _RPA_ACCENT, _COLOR_OK, _COLOR_WARN, _COLOR_ERR,
        _html_open, _html_close, _to_israel_time,
    )
except ImportError:
    _RPA_BLUE = "#1e3a5f"
    _RPA_ACCENT = "#2471a3"
    _COLOR_OK = "#27ae60"
    _COLOR_WARN = "#f39c12"
    _COLOR_ERR = "#e74c3c"
    _html_open = None
    _html_close = None
    _to_israel_time = None

# =============================================================================
# SESSION 27: Per-Run Cost Accumulator
# =============================================================================

class _CostTracker:
    """Accumulates AI costs per classification run. Thread-safe reset per call."""
    def __init__(self):
        self.reset()

    def reset(self):
        self._costs = {}  # model_name -> total_cost
        self._calls = {}  # model_name -> call_count

    def add(self, model, cost):
        self._costs[model] = self._costs.get(model, 0) + cost
        self._calls[model] = self._calls.get(model, 0) + 1

    def total(self):
        return sum(self._costs.values())

    def summary(self):
        if not self._costs:
            return ""
        lines = ["ğŸ’° Cost Summary:"]
        for model in sorted(self._costs.keys()):
            lines.append(f"    {model}: {self._calls[model]} calls = ${self._costs[model]:.4f}")
        lines.append(f"    TOTAL: ${self.total():.4f}")
        return "\n".join(lines)

_cost_tracker = _CostTracker()

# =============================================================================
# HS CODE VALIDATION HELPERS
# =============================================================================

_HS_RE = re.compile(r'^\d{4,10}$')

def _is_valid_hs(code):
    """Check if a string looks like a real HS code (4-10 digits, chapter 01-97)."""
    if not code or not isinstance(code, str):
        return False
    clean = code.replace(".", "").replace("/", "").replace(" ", "").strip()
    if not _HS_RE.match(clean):
        return False
    try:
        chapter = int(clean[:2])
        return 1 <= chapter <= 97
    except (ValueError, IndexError):
        return False


def _is_product_description(text):
    """Check if text looks like a real product description, not raw email/HTML."""
    if not text or not isinstance(text, str):
        return False
    if text.startswith("=== "):
        return False
    if "&nbsp;" in text or "<br" in text.lower() or "<div" in text.lower():
        return False
    if len(text) > 400 and text.count("===") > 1:
        return False
    return True


# â”€â”€ KNOWN_FAILURES pattern matching â”€â”€
# Programmatic guard against historically known classification mistakes.
# Maps failure ID â†’ (product_keywords, wrong_heading_prefix, correct_heading, warning_he)
_KNOWN_FAILURE_PATTERNS = [
    {
        "id": "F001",
        "product_keywords": {"kiwi", "×§×™×•×•×™", "×§×•×•×™"},
        "wrong_headings": {"0305", "1604", "1605"},  # fish/caviar chapters
        "correct_hint": "08.05 / 08.10 (fruits)",
        "warning": "âš ï¸ F001: ×§×™×•×•×™ = ×¤×¨×™ (×¤×¨×§ 8), ×œ× ×§×•×•×™××¨ (×¤×¨×§ 3/16)",
    },
    {
        "id": "F003",
        "product_keywords": {"tire", "tyre", "tires", "tyres", "pneumatic", "×¦××™×’", "×¦××™×’×™×"},
        "wrong_headings": {"4001", "4002", "4003", "4004", "4005"},  # raw rubber headings
        "correct_hint": "40.11 (pneumatic tires)",
        "warning": "âš ï¸ F003: ×¦××™×’×™× = 40.11 (×¦××™×’×™× ×¤× ××•××˜×™×™×), ×œ× 40.01 (×’×•××™ ×’×•×œ××™)",
    },
]


def _check_known_failure_patterns(classifications):
    """Check classification results against KNOWN_FAILURES patterns.

    Adds warning to any classification that matches a known failure pattern.
    Returns the number of warnings added.
    """
    warnings_added = 0
    for c in classifications:
        hs = c.get("hs_code", "").replace(".", "").replace("/", "").replace(" ", "")
        if not hs or len(hs) < 4:
            continue
        heading = hs[:4]
        desc = (c.get("description", "") or c.get("item", "") or "").lower()
        if not desc:
            continue

        for pattern in _KNOWN_FAILURE_PATTERNS:
            if heading not in pattern["wrong_headings"]:
                continue
            # Check if any product keyword appears in description
            if any(kw in desc for kw in pattern["product_keywords"]):
                c["known_failure_warning"] = pattern["warning"]
                c["known_failure_id"] = pattern["id"]
                c["known_failure_correct_hint"] = pattern["correct_hint"]
                if c.get("confidence") == "×’×‘×•×”×”":
                    c["confidence"] = "× ××•×›×”"
                elif c.get("confidence") == "×‘×™× ×•× ×™×ª":
                    c["confidence"] = "× ××•×›×”"
                print(f"    ğŸš¨ KNOWN FAILURE {pattern['id']}: {desc[:60]} â†’ HS {hs} matches failure pattern!")
                warnings_added += 1
                break
    return warnings_added


# =============================================================================
# SESSION 11: Tracking Code & Subject Line Builder
# =============================================================================

def generate_tracking_code():
    """Generate unique RCB tracking code: RCB-YYYYMMDD-XXXXX"""
    date_part = datetime.now().strftime("%Y%m%d")
    random_part = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
    return f"RCB-{date_part}-{random_part}"


def build_rcb_subject(invoice_data, status="ACK", invoice_score=None):
    """
    Build standardized RCB email subject line in English.
    
    Format: [TRACKING] Direction | RPA: XXX | Seller: XXX | Buyer: XXX | Freight | BL/AWB | STATUS
    
    Args:
        invoice_data: Dict from Agent 1 extraction
        status: "ACK" (acknowledged), "FINAL" (complete), "CLARIFICATION" (needs info)
        invoice_score: Optional invoice validation score
        
    Returns:
        Formatted subject line string
    """
    # Generate tracking code
    tracking = generate_tracking_code()
    
    # Direction
    direction = invoice_data.get('direction', 'unknown')
    if direction == 'import':
        direction_text = "Import to IL"
    elif direction == 'export':
        direction_text = "Export from IL"
    else:
        direction_text = "Direction TBD"
    
    # RPA file number
    rpa_file = invoice_data.get('rpa_file_number', '')
    rpa_text = f"RPA:{rpa_file}" if rpa_file else "RPA:TBD"
    
    # Seller (truncate to 20 chars)
    seller = invoice_data.get('seller', '')
    if seller:
        # Extract company name (first part before comma or address)
        seller = seller.split(',')[0].strip()[:20]
    seller_text = f"S:{seller}" if seller else "S:Unknown"
    
    # Buyer (truncate to 20 chars)
    buyer = invoice_data.get('buyer', '')
    if buyer:
        buyer = buyer.split(',')[0].strip()[:20]
    buyer_text = f"B:{buyer}" if buyer else "B:Unknown"
    
    # Freight type
    freight_type = invoice_data.get('freight_type', 'unknown')
    if freight_type == 'sea':
        freight_text = "SEA"
    elif freight_type == 'air':
        freight_text = "AIR"
    else:
        freight_text = "FRT:TBD"
    
    # BL or AWB number
    bl_number = invoice_data.get('bl_number', '')
    awb_number = invoice_data.get('awb_number', '')
    if bl_number:
        transport_text = f"BL:{bl_number[:15]}"
    elif awb_number:
        transport_text = f"AWB:{awb_number[:15]}"
    else:
        transport_text = "REF:TBD"
    
    # Status with score if available
    if status == "CLARIFICATION" or (invoice_score is not None and invoice_score < 70):
        status_text = "âš ï¸CLARIFICATION"
    elif status == "FINAL" or (invoice_score is not None and invoice_score >= 70):
        status_text = "âœ…FINAL"
    else:
        status_text = "ğŸ“¥ACK"
    
    # Build subject
    subject = f"[{tracking}] {direction_text} | {rpa_text} | {seller_text} | {buyer_text} | {freight_text} | {transport_text} | {status_text}"
    
    return subject, tracking


def extract_tracking_from_subject(subject):
    """Extract RCB tracking code from existing subject if present"""
    import re
    match = re.search(r'\[RCB-\d{8}-[A-Z0-9]{5}\]', subject)
    if match:
        return match.group(0)[1:-1]  # Remove brackets
    return None


def clean_firestore_data(data):
    """Convert Firestore timestamps to strings for JSON serialization"""
    if isinstance(data, dict):
        return {k: clean_firestore_data(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [clean_firestore_data(i) for i in data]
    elif hasattr(data, 'isoformat'):
        return data.isoformat()
    return data


def call_claude(api_key, system_prompt, user_prompt, max_tokens=2000):
    """Call Claude API - Sonnet 4.5 (Session 15: upgraded from Sonnet 4)
    Session 26: Added cost tracking
    Session 48: Added None/empty key guard"""
    if not api_key:
        print("    âŒ call_claude: api_key is None/empty â€” cannot call Claude API")
        return None
    try:
        response = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": api_key.strip(),
                "content-type": "application/json",
                "anthropic-version": "2023-06-01"
            },
            json={
                "model": "claude-sonnet-4-20250514",
                "max_tokens": max_tokens,
                "system": system_prompt,
                "messages": [{"role": "user", "content": user_prompt}]
            },
            timeout=120
        )
        if response.status_code == 200:
            data = response.json()
            # Cost tracking: Claude Sonnet 4.5 pricing ($3/$15 per MTok)
            if COST_TRACKING_ENABLED:
                usage = data.get("usage", {})
                inp_tok = usage.get("input_tokens", 0)
                out_tok = usage.get("output_tokens", 0)
                cost = (inp_tok * 3.0 + out_tok * 15.0) / 1_000_000
                print(f"    ğŸ’° Claude: {inp_tok}+{out_tok} tokens = ${cost:.4f}")
                _cost_tracker.add("Claude Sonnet", cost)
            return data['content'][0]['text']
        else:
            print(f"Claude API error: {response.status_code} - {response.text[:200]}")
            return None
    except Exception as e:
        print(f"Claude call error: {e}")
        return None


# =============================================================================
# SESSION 15: Gemini API Integration (cost optimization)
# =============================================================================

# =============================================================================
# SESSION 27: ChatGPT / OpenAI Integration (cross-check)
# =============================================================================

def call_chatgpt(openai_key, system_prompt, user_prompt, max_tokens=2000, model="gpt-4o-mini"):
    """Call OpenAI ChatGPT API.

    Session 27: Added for three-way cross-check.
    Models:
      - gpt-4o-mini: Fast, cheap (~$0.15/$0.60 per MTok)
      - gpt-4o: Higher quality (~$2.50/$10 per MTok)
    """
    if not openai_key:
        return None
    try:
        import openai
        client = openai.OpenAI(api_key=openai_key)
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            max_tokens=max_tokens,
            temperature=0.3,
        )
        result = response.choices[0].message.content
        # Cost tracking
        if COST_TRACKING_ENABLED:
            usage = response.usage
            inp_tok = usage.prompt_tokens if usage else 0
            out_tok = usage.completion_tokens if usage else 0
            if "mini" in model:
                cost = (inp_tok * 0.15 + out_tok * 0.60) / 1_000_000
            else:  # gpt-4o
                cost = (inp_tok * 2.50 + out_tok * 10.0) / 1_000_000
            print(f"    ğŸ’° ChatGPT ({model}): {inp_tok}+{out_tok} tokens = ${cost:.4f}")
            _cost_tracker.add(f"ChatGPT {model}", cost)
        return result
    except ImportError:
        print("    âš ï¸ openai package not installed, ChatGPT unavailable")
        return None
    except Exception as e:
        print(f"    ChatGPT call error ({model}): {e}")
        return None


def call_gemini(gemini_key, system_prompt, user_prompt, max_tokens=2000, model="gemini-2.5-flash"):
    """Call Google Gemini API

    Session 15: Added for cost optimization.
    Session 48: Added 429 fast-fail â€” skip all Gemini after quota exhaustion.
    Models:
      - gemini-2.5-flash: Simple tasks (extraction, regulatory, FTA, risk) ~$0.15/$0.60 per MTok
      - gemini-2.5-pro: Medium tasks (synthesis) ~$1.25/$10 per MTok

    Falls back to Claude if Gemini fails or key is missing.
    """
    global _gemini_quota_exhausted_at
    if not gemini_key:
        return None
    # Session 48+CRIT-2: Skip Gemini if 429 within last 60s
    if _gemini_quota_exhausted_at and (time.time() - _gemini_quota_exhausted_at < 60):
        return None
    try:
        response = requests.post(
            f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={gemini_key}",
            headers={"content-type": "application/json"},
            json={
                "systemInstruction": {"parts": [{"text": system_prompt}]},
                "contents": [{"parts": [{"text": user_prompt}]}],
                "generationConfig": {
                    "maxOutputTokens": max_tokens,
                    "temperature": 0.3
                }
            },
            timeout=120
        )
        if response.status_code == 200:
            data = response.json()
            # Cost tracking: Gemini pricing
            if COST_TRACKING_ENABLED:
                usage = data.get("usageMetadata", {})
                inp_tok = usage.get("promptTokenCount", 0)
                out_tok = usage.get("candidatesTokenCount", 0)
                if "flash" in model:
                    cost = (inp_tok * 0.15 + out_tok * 0.60) / 1_000_000
                else:  # pro
                    cost = (inp_tok * 1.25 + out_tok * 10.0) / 1_000_000
                print(f"    ğŸ’° Gemini ({model}): {inp_tok}+{out_tok} tokens = ${cost:.4f}")
                _cost_tracker.add(f"Gemini {model}", cost)
            candidates = data.get("candidates", [])
            if candidates:
                finish_reason = candidates[0].get("finishReason", "UNKNOWN")
                if finish_reason not in ("STOP", "UNKNOWN"):
                    print(f"    âš ï¸ Gemini finishReason: {finish_reason} (model={model})")
                parts = candidates[0].get("content", {}).get("parts", [])
                if parts:
                    text = parts[0].get("text", "")
                    # Strip markdown code fences that Gemini often wraps around JSON
                    if text.startswith("```"):
                        text = text.split("\n", 1)[1] if "\n" in text else text[3:]
                    if text.endswith("```"):
                        text = text[:-3]
                    return text.strip()
        else:
            # Session 48: Detect 429 quota exhaustion â€” fast-fail for rest of run
            if response.status_code == 429:
                _gemini_quota_exhausted_at = time.time()
                print(f"    âš¡ Gemini 429 quota exhausted â€” skipping Gemini for 60s")
            else:
                print(f"Gemini API error ({model}): {response.status_code} - {response.text[:200]}")
            return None
    except Exception as e:
        print(f"Gemini call error ({model}): {e}")
        return None


def call_gemini_fast(gemini_key, system_prompt, user_prompt, max_tokens=2000):
    """Gemini 2.5 Flash - for simple structured tasks (Agents 1,3,4,5)"""
    return call_gemini(gemini_key, system_prompt, user_prompt, max_tokens, "gemini-2.5-flash")


def call_gemini_pro(gemini_key, system_prompt, user_prompt, max_tokens=2000):
    """Gemini 2.5 Pro - for medium complexity tasks (Agent 6 synthesis)"""
    return call_gemini(gemini_key, system_prompt, user_prompt, max_tokens, "gemini-2.5-pro")


def call_ai(api_key, gemini_key, system_prompt, user_prompt, max_tokens=2000, tier="fast"):
    """Smart router: picks the right model based on task tier.
    
    Session 15: Central routing function with automatic fallback.
    - tier="smart": Claude Sonnet 4.5 (Agent 2 - classification)
    - tier="pro": Gemini 2.5 Pro (Agent 6 - synthesis)  
    - tier="fast": Gemini 2.5 Flash (Agents 1,3,4,5 - simple tasks)
    
    If Gemini fails or key is missing, falls back to Claude automatically.
    """
    result = None
    
    if tier == "smart":
        # Always use Claude for the hard classification task
        result = call_claude(api_key, system_prompt, user_prompt, max_tokens)
    elif tier == "pro":
        # Try Gemini Pro first, fallback to Claude
        result = call_gemini_pro(gemini_key, system_prompt, user_prompt, max_tokens)
        if not result:
            print("    â†©ï¸ Gemini Pro fallback â†’ Claude")
            result = call_claude(api_key, system_prompt, user_prompt, max_tokens)
    else:  # "fast"
        # Try Gemini Flash first, fallback to Claude
        result = call_gemini_fast(gemini_key, system_prompt, user_prompt, max_tokens)
        if not result:
            print("    â†©ï¸ Gemini Flash fallback â†’ Claude")
            result = call_claude(api_key, system_prompt, user_prompt, max_tokens)
    
    return result

def _build_elimination_context(elimination_results):
    """Build context string from elimination results for Agent 2.

    Summarizes survivors, eliminated codes, and key elimination steps so that
    Agent 2 can focus on the narrowed candidate set.
    """
    lines = ["\n\n=== ELIMINATION ENGINE RESULTS ==="]
    for desc_key, elim in elimination_results.items():
        survivors = elim.get("survivors", [])
        eliminated = elim.get("eliminated", [])
        steps = elim.get("steps", [])
        lines.append(f"\nProduct: {desc_key}")
        if survivors:
            surv_codes = [s["hs_code"] for s in survivors]
            lines.append(f"  Surviving candidates: {', '.join(surv_codes)}")
            for s in survivors:
                conf = s.get("confidence", 0)
                desc = s.get("description", "")[:80]
                lines.append(f"    {s['hs_code']} (confidence {conf}): {desc}")
        if eliminated:
            elim_codes = [e["hs_code"] for e in eliminated]
            lines.append(f"  Eliminated: {', '.join(elim_codes)}")
        # Include key elimination steps (up to 5 most impactful)
        impactful = [st for st in steps if st.get("eliminated_codes")]
        for st in impactful[:5]:
            codes = ", ".join(st.get("eliminated_codes", []))
            lines.append(f"  Step [{st.get('level', '')}]: {st.get('reasoning', '')[:120]} â†’ eliminated {codes}")
        challenges = elim.get("challenges", [])
        if challenges:
            lines.append(f"  Devil's advocate challenges: {len(challenges)}")
            for ch in challenges[:2]:
                lines.append(f"    {ch.get('candidate', '')}: {ch.get('counter_argument', '')[:100]}")
    lines.append("=== END ELIMINATION ===\n")
    return "\n".join(lines)


def query_tariff(db, search_terms):
    """Legacy stub â€” tool engine search_tariff (tool #2) provides indexed tariff
    lookups via intelligence.pre_classify(). This previously streamed all 11,753
    tariff docs per call. Agent 2 prompt uses tariff_data[:15] for examples only;
    the tool loop already provides comprehensive tariff context."""
    return []

def query_ministry_index(db):
    """Get ministry requirements"""
    results = []
    try:
        docs = db.collection('ministry_index').stream()
        for doc in docs:
            results.append(doc.to_dict())
    except Exception as e:
        print(f"Ministry query error: {e}")
    return clean_firestore_data(results)

def query_classification_rules(db):
    """Get classification rules"""
    rules = []
    try:
        docs = db.collection('classification_rules').stream()
        for doc in docs:
            rules.append(doc.to_dict())
    except Exception as e:
        print(f"Rules query error: {e}")
    return rules

def _try_parse_agent1(result, model_name):
    """Try to parse Agent 1 JSON response. Returns parsed dict or None."""
    try:
        if result:
            start, end = result.find('{'), result.rfind('}') + 1
            if start != -1 and end > start:
                parsed = json.loads(result[start:end])
                items = parsed.get("items", [])
                print(f"    ğŸ“¦ Agent 1 ({model_name}): {len(items)} items extracted, keys={list(parsed.keys())}")
                return parsed
            else:
                print(f"    âš ï¸ Agent 1 ({model_name}): No complete JSON in response ({len(result)} chars): {result[:200]}")
        else:
            print(f"    âš ï¸ Agent 1 ({model_name}): returned None/empty")
    except Exception as e:
        print(f"    âš ï¸ Agent 1 ({model_name}) JSON parse error: {e}")
        print(f"    âš ï¸ Agent 1 ({model_name}) raw ({len(result)} chars): {result[:300]}")
    return None


def run_document_agent(api_key, doc_text, gemini_key=None, openai_key=None):
    """Agent 1: Extract invoice data - Updated Session 11 with shipping details
    Session 15: Uses Gemini Flash (simple extraction task)
    Session 48: Triple fallback chain â€” Gemini â†’ Claude â†’ ChatGPT"""
    system = """××ª×” ×¡×•×›×Ÿ ×—×™×œ×•×¥ ××™×“×¢ ×××¡××›×™ ×¡×—×¨ ×‘×™× ×œ××•××™.
×”××¡××š ×™×›×•×œ ×œ×”×™×•×ª: ×—×©×‘×•× ×™×ª ××¡×—×¨×™×ª, ×—×©×‘×•×Ÿ ×¤×¨×•×¤×•×¨××”, ×”×¦×¢×ª ××—×™×¨, ×¨×©×™××ª ××—×™×¨×™×, ××• ×›×œ ××¡××š ×”××›×™×œ ×¤×¨×™×˜×™× ×œ×¡×™×•×•×’ ××›×¡.
×—×œ×¥ ××”××¡××š JSON ×¢× ×›×œ ×”×©×“×•×ª ×”×‘××™×:
{
    "seller": "",
    "buyer": "",
    "items": [{"description":"","quantity":"","unit_price":"","total":"","origin_country":""}],
    "invoice_number": "",
    "invoice_date": "",
    "total_value": "",
    "currency": "",
    "incoterms": "",
    "direction": "import/export/unknown",
    "freight_type": "sea/air/unknown",
    "bl_number": "",
    "awb_number": "",
    "rpa_file_number": "",
    "vessel_name": "",
    "port_of_loading": "",
    "port_of_discharge": ""
}

×”× ×—×™×•×ª:
- CRITICAL: Extract EACH product as a SEPARATE item in the items array. Every line item on the invoice must be its own entry. If the invoice has 10 products, return 10 items. Include the product name, quantity, unit price, total, and origin country for each.
- ×× ×”××¡××š ×”×•× ×”×¦×¢×ª ××—×™×¨ ××• ×¤×¨×•×¤×•×¨××” â€” ×—×œ×¥ ××ª ×”×¤×¨×™×˜×™× ×‘×“×™×•×§ ×›××• ××—×©×‘×•× ×™×ª. ×”×©×ª××© ×‘××¡×¤×¨ ×”×”×¦×¢×” ×›-invoice_number.
- ×›×œ ×©×•×¨×” ×‘×˜×‘×œ×ª ××—×™×¨×™×/×¤×¨×™×˜×™× = item × ×¤×¨×“. ××œ ×ª×“×œ×’ ×¢×œ ×©×•×¨×•×ª.
- Chinese product names (ä¸­æ–‡) must be translated to English in the description field.
- Do NOT combine multiple products into one item. Do NOT summarize.
- direction: "import" ×× ×”×¡×—×•×¨×” ××’×™×¢×” ×œ×™×©×¨××œ, "export" ×× ×™×•×¦××ª ××™×©×¨××œ
- freight_type: "sea" ×× ×™×© B/L ××• ××•× ×™×™×”, "air" ×× ×™×© AWB ××• ×˜×™×¡×”
- bl_number: ××¡×¤×¨ ×©×˜×¨ ××˜×¢×Ÿ ×™××™ (Bill of Lading)
- awb_number: ××¡×¤×¨ ×©×˜×¨ ××˜×¢×Ÿ ××•×•×™×¨×™ (Air Waybill)
- rpa_file_number: ××¡×¤×¨ ×ª×™×§ RPA ×× ××•×¤×™×¢

JSON ×‘×œ×‘×“."""
    # Fallback 1: Gemini Flash (cheapest) â€” via call_ai which auto-falls back to Claude
    result = call_ai(api_key, gemini_key, system, doc_text[:6000], max_tokens=4096, tier="fast")
    parsed = _try_parse_agent1(result, "call_ai(Geminiâ†’Claude)")
    if parsed:
        return parsed

    # Fallback 2: Claude direct (in case call_ai's fallback also failed)
    print("    ğŸ”„ Agent 1: call_ai failed, retrying Claude directly...")
    result = call_claude(api_key, system, doc_text[:6000], max_tokens=4096)
    parsed = _try_parse_agent1(result, "Claude-direct")
    if parsed:
        return parsed

    # Fallback 3: ChatGPT (Session 48 â€” third fallback, never leave pipeline with 0 models)
    if openai_key:
        print("    ğŸ”„ Agent 1: Claude failed, trying ChatGPT (gpt-4o-mini)...")
        result = call_chatgpt(openai_key, system, doc_text[:6000], max_tokens=4096, model="gpt-4o-mini")
        parsed = _try_parse_agent1(result, "ChatGPT")
        if parsed:
            return parsed

    print("    âš ï¸ Agent 1: ALL models failed â€” returning doc_text[:500] as single item")
    return {"items": [{"description": doc_text[:500]}]}


def run_classification_agent(api_key, items, tariff_data, rules, knowledge_context="", gemini_key=None):
    """Agent 2: Classify items using Israeli HS codes
    Session 15: STAYS on Claude Sonnet 4.5 (highest quality for core task)
    Session 51: Legal context from customs_law.py injected BEFORE knowledge context"""
    # Session 51: Inject embedded customs law expertise at the TOP
    _legal_ctx = ""
    if CUSTOMS_LAW_AVAILABLE:
        try:
            _legal_ctx = format_legal_context_for_prompt()
        except Exception:
            pass
    system = f"""{_legal_ctx}××ª×” ×¡×•×›×Ÿ ×¡×™×•×•×’ ××›×¡ ×™×©×¨××œ×™ ××•××—×”.
××•× ×—×™×: ×¢××™×œ ××›×¡ / ×¡×•×›×Ÿ ××›×¡ â€” ×œ×¢×•×œ× ×œ× ××ª×•×•×š ××›×¡.

{knowledge_context}

×›×œ×œ×™×:
{json.dumps(rules[:10], ensure_ascii=False)}

×ª×¢×¨×™×¤×•×Ÿ (×“×•×’×××•×ª):
{json.dumps(tariff_data[:15], ensure_ascii=False)}

×¡×•×•×’ ×›×œ ×¤×¨×™×˜ ×¢× ×§×•×“ HS ×™×©×¨××œ×™ (8-10 ×¡×¤×¨×•×ª).
×¤×œ×˜ JSON:
{{"classifications":[{{"item":"","hs_code":"","duty_rate":"","confidence":"×’×‘×•×”×”/×‘×™× ×•× ×™×ª/× ××•×›×”","reasoning":""}}]}}"""
    
    result = call_ai(api_key, gemini_key, system, f"×¤×¨×™×˜×™× ×œ×¡×™×•×•×’:\n{json.dumps(items, ensure_ascii=False)}", 3000, tier="smart")
    try:
        if result:
            start, end = result.find('{'), result.rfind('}') + 1
            if start != -1: return json.loads(result[start:end])
    except Exception as e:
        print(f"    âŒ Agent 2 (Classification) JSON parse error: {e} | raw[:200]={result[:200] if result else 'None'}")
    return {"classifications": []}


def run_regulatory_agent(api_key, classifications, ministry_data, gemini_key=None):
    """Agent 3: Check regulatory requirements
    Session 15: Uses Gemini Flash (structured matching task)"""
    system = f"""××ª×” ×¡×•×›×Ÿ ×¨×’×•×œ×¦×™×”. ×‘×“×•×§ ××™×œ×• ××™×©×•×¨×™× × ×“×¨×©×™× ×œ×¤×™ ×”×¡×™×•×•×’×™×.

××©×¨×“×™× ×•×“×¨×™×©×•×ª:
{json.dumps(ministry_data[:20], ensure_ascii=False)}

×¤×œ×˜ JSON:
{{"regulatory":[{{"hs_code":"","ministries":[{{"name":"","required":true/false,"regulation":""}}]}}]}}"""
    
    result = call_ai(api_key, gemini_key, system, f"×¡×™×•×•×’×™×:\n{json.dumps(classifications, ensure_ascii=False)}", tier="fast")
    try:
        if result:
            start, end = result.find('{'), result.rfind('}') + 1
            if start != -1: return json.loads(result[start:end])
    except Exception as e:
        print(f"    âŒ Agent 3 (Regulatory) JSON parse error: {e} | raw[:200]={result[:200] if result else 'None'}")
    return {"regulatory": []}


def run_fta_agent(api_key, classifications, origin_country, gemini_key=None):
    """Agent 4: Check FTA eligibility
    Session 15: Uses Gemini Flash (country-agreement matching)"""
    system = """××ª×” ×¡×•×›×Ÿ ×”×¡×›××™ ×¡×—×¨. ×‘×“×•×§ ×–×›××•×ª ×œ×”×¢×“×¤×•×ª ××›×¡.

×”×¡×›××™× ×¤×¢×™×œ×™×: EU, USA, UK, EFTA, Turkey, Jordan, Egypt, Mercosur, Mexico, Canada

×¤×œ×˜ JSON:
{"fta":[{"hs_code":"","country":"","agreement":"","eligible":true/false,"preferential":"","documents_needed":""}]}"""
    
    result = call_ai(api_key, gemini_key, system, f"×¡×™×•×•×’×™×: {json.dumps(classifications, ensure_ascii=False)}\n××¨×¥ ××§×•×¨: {origin_country}", tier="fast")
    try:
        if result:
            start, end = result.find('{'), result.rfind('}') + 1
            if start != -1: return json.loads(result[start:end])
    except Exception as e:
        print(f"    âŒ Agent 4 (FTA) JSON parse error: {e} | raw[:200]={result[:200] if result else 'None'}")
    return {"fta": []}


def run_risk_agent(api_key, invoice_data, classifications, gemini_key=None):
    """Agent 5: Risk assessment
    Session 15: Uses Gemini Flash (pattern detection task)"""
    system = """××ª×” ×¡×•×›×Ÿ ×”×¢×¨×›×ª ×¡×™×›×•× ×™×. ×‘×“×•×§:
1. ×¢×¨×š × ××•×š ×—×©×•×“
2. ×¡×™×•×•×’ ×©×’×•×™ ××¤×©×¨×™
3. ××§×•×¨ ×‘×¢×™×™×ª×™
4. ×—×•×¡×¨ ×”×ª×××”

×¤×œ×˜ JSON:
{"risk":{"level":"× ××•×š/×‘×™× ×•× ×™/×’×‘×•×”","items":[{"item":"","issue":"","recommendation":""}]}}"""
    
    result = call_ai(api_key, gemini_key, system, f"×—×©×‘×•× ×™×ª: {json.dumps(invoice_data, ensure_ascii=False)}\n×¡×™×•×•×’×™×: {json.dumps(classifications, ensure_ascii=False)}", tier="fast")
    try:
        if result:
            start, end = result.find('{'), result.rfind('}') + 1
            if start != -1: return json.loads(result[start:end])
    except Exception as e:
        print(f"    âŒ Agent 5 (Risk) JSON parse error: {e} | raw[:200]={result[:200] if result else 'None'}")
    return {"risk": {"level": "× ××•×š", "items": []}}


def run_synthesis_agent(api_key, all_results, gemini_key=None):
    """Agent 6: Final synthesis with proper Hebrew formatting
    Session 15: Uses Gemini 2.5 Pro (good Hebrew, lower cost than Claude)"""
    system = """××ª×” ×¡×•×›×Ÿ ×¡×™×›×•× ××§×¦×•×¢×™. ×›×ª×•×‘ ×¡×™×›×•× ××¢×•×¦×‘ ×”×™×˜×‘ ×‘×¢×‘×¨×™×ª.
××•× ×—×™×: ×¢××™×œ ××›×¡ / ×¡×•×›×Ÿ ××›×¡ â€” ×œ×¢×•×œ× ×œ× ××ª×•×•×š ××›×¡.

×—×•×‘×” ×œ×¢×§×•×‘ ××—×¨ ×”×¤×•×¨××˜ ×”×‘× ×‘×“×™×•×§:

ğŸ“‹ ××” × ××¦×:
[×›×ª×•×‘ 2-3 ××©×¤×˜×™× ×”××ª××¨×™× ××ª ×ª×•×›×Ÿ ×”×—×©×‘×•× ×™×ª, ×”××•×¦×¨×™×, ××¨×¥ ×”××§×•×¨, ×•×”×¢×¨×š. ×›×œ ××©×¤×˜ ×‘×©×•×¨×” ×—×“×©×”.]

ğŸ“Œ ×”××œ×¦×•×ª ×¢×™×§×¨×™×•×ª:
â€¢ [×”××œ×¦×” ×¨××©×•× ×”]
â€¢ [×”××œ×¦×” ×©× ×™×™×”]
â€¢ [×”××œ×¦×” ×©×œ×™×©×™×ª ×× ×¨×œ×•×•× ×˜×™]

âš ï¸ ××–×”×¨×•×ª:
â€¢ [××–×”×¨×” ×¨××©×•× ×” ×× ×™×©]
â€¢ [××–×”×¨×” ×©× ×™×™×” ×× ×™×©]

×›×œ×œ×™ ×¢×™×¦×•×‘ ×—×©×•×‘×™×:
- ×”×©×ª××© ×‘×©×•×¨×” ×¨×™×§×” ×‘×™×Ÿ ×›×œ ×§×˜×¢
- ×”×©×ª××© ×‘× ×§×•×“×•×ª (â€¢) ×œ×¨×©×™××•×ª
- ×¡×™×™× ×›×œ ××©×¤×˜ ×‘× ×§×•×“×”
- ××œ ×ª×©×ª××© ×‘×›×•×›×‘×™×•×ª (**) ××• ×‘×¡×™×× ×™ ×¢×™×¦×•×‘ ××—×¨×™×
- ×›×ª×•×‘ ×‘×¢×‘×¨×™×ª ×ª×§× ×™×ª ×•×‘×¨×•×¨×”
- ×× ××™×Ÿ ××–×”×¨×•×ª, ×”×©××˜ ××ª ×”×§×˜×¢ ×”×–×” ×œ×—×œ×•×˜×™×Ÿ

×˜×§×¡×˜ ××¢×•×¦×‘ ×‘×œ×‘×“, ×œ× JSON."""
    
    return call_ai(api_key, gemini_key, system, json.dumps(all_results, ensure_ascii=False)[:4000], tier="pro") or "×œ× × ×™×ª×Ÿ ×œ×™×™×¦×¨ ×¡×™×›×•×."


def _link_invoice_to_classifications(items, classifications):
    """Link each classification back to its source invoice line item.

    Uses simple substring matching on descriptions. Adds line_number,
    invoice_description, quantity, unit_price, total, origin_country
    to each classification dict.
    """
    if not items or not classifications:
        return classifications

    for c in classifications:
        c_desc = (c.get("item") or "").lower().strip()
        if not c_desc:
            continue
        best_idx = -1
        best_score = 0
        for idx, inv_item in enumerate(items):
            if not isinstance(inv_item, dict):
                continue
            inv_desc = (inv_item.get("description") or "").lower().strip()
            if not inv_desc:
                continue
            # Simple overlap: check if key words match
            if inv_desc in c_desc or c_desc in inv_desc:
                score = len(inv_desc)
                if score > best_score:
                    best_score = score
                    best_idx = idx
            else:
                # Fallback: count shared words
                c_words = set(c_desc.split())
                inv_words = set(inv_desc.split())
                overlap = len(c_words & inv_words)
                if overlap > best_score:
                    best_score = overlap
                    best_idx = idx

        if best_idx >= 0:
            inv = items[best_idx]
            c["line_number"] = best_idx + 1
            c["invoice_description"] = inv.get("description", "")
            c["quantity"] = inv.get("quantity", "")
            c["unit_price"] = inv.get("unit_price", "")
            c["total_value"] = inv.get("total", "")
            c["origin_country"] = inv.get("origin_country", "")
        else:
            c["line_number"] = 0

    return classifications


def run_full_classification(api_key, doc_text, db, gemini_key=None, openai_key=None):
    """Run complete multi-agent classification
    Session 15: Now accepts gemini_key for cost-optimized multi-model routing
    Session 18: Intelligence module runs BEFORE AI agents"""
    try:
        # Agent 1: Extract (Gemini Flash)
        print("    ğŸ” Agent 1: Extracting... [Gemini Flash]")
        invoice = run_document_agent(api_key, doc_text, gemini_key=gemini_key, openai_key=openai_key)
        if not isinstance(invoice, dict):
            print(f"    âš ï¸ Agent 1 returned non-dict: {type(invoice)} â€” using fallback")
            invoice = {"items": [{"description": doc_text[:500]}]}
        items = invoice.get("items") or [{"description": doc_text[:500]}]
        if not isinstance(items, list):
            print(f"    âš ï¸ Agent 1 items is not a list: {type(items)} â€” using fallback")
            items = [{"description": doc_text[:500]}]

        # DEBUG: Log exactly what Agent 1 extracted
        print(f"    ğŸ“¦ Agent 1 extracted {len(items)} items:")
        for idx, item in enumerate(items[:10]):
            if isinstance(item, dict):
                desc = item.get('description', '')[:80]
                qty = item.get('quantity', '')
                origin_c = item.get('origin_country', '')
                print(f"       [{idx+1}] {desc} | qty={qty} | origin={origin_c}")
            else:
                print(f"       [{idx+1}] NOT A DICT: {str(item)[:80]}")

        origin = items[0].get("origin_country", "") if items and isinstance(items[0], dict) else ""

        # â”€â”€ DOCUMENT PARSER: Identify each document and extract structured fields â”€â”€
        parsed_documents = []
        if DOCUMENT_PARSER_AVAILABLE:
            try:
                parsed_documents = parse_all_documents(doc_text)
                if parsed_documents:
                    types_found = [d["type_info"]["name_en"] for d in parsed_documents]
                    print(f"    ğŸ“‘ Document Parser: {len(parsed_documents)} docs â€” {', '.join(types_found)}")
                    for pd in parsed_documents:
                        comp = pd.get("completeness", {})
                        if comp.get("missing"):
                            crit = [m["name_he"] for m in comp["missing"] if m["importance"] == "critical"]
                            if crit:
                                print(f"       âš ï¸ {pd.get('filename','')}: missing critical â€” {', '.join(crit)}")
            except Exception as e:
                print(f"    âš ï¸ Document parser error: {e}")

        # â”€â”€ SHIPMENT TRACKER: Feed BL/DO/invoice fields to tracker â”€â”€
        tracker_info = None
        if TRACKER_AVAILABLE and parsed_documents:
            try:
                bl_num = invoice.get("bl_number", "") or ""
                shipment_id = bl_num or invoice.get("awb_number", "") or "PENDING"
                direction_val = invoice.get("direction", "import")
                transport = "sea_fcl" if invoice.get("freight_type") == "sea" else (
                    "air" if invoice.get("freight_type") == "air" else None)

                tracker = create_tracker(
                    shipment_id=shipment_id,
                    direction=direction_val,
                    transport_mode=transport,
                )
                feed_parsed_documents(tracker, parsed_documents, invoice_data=invoice)

                # Also add invoice doc if Agent 1 extracted it
                if invoice.get("invoice_number") or invoice.get("seller"):
                    tracker.add_document(Document(
                        doc_type=TrackerDocType.INVOICE,
                        doc_number=invoice.get("invoice_number", ""),
                    ))

                tracker_info = tracker.to_dict()
                print(f"    ğŸ“¦ Tracker: phase={tracker.phase_hebrew}, "
                      f"docs={len(tracker.documents)}, "
                      f"missing={len(tracker.missing_docs)}")
            except Exception as e:
                print(f"    âš ï¸ Tracker error: {e}")

        # â”€â”€ PRE-CLASSIFY: System's own brain BEFORE AI â”€â”€
        intelligence_context = ""
        intelligence_results = {}
        doc_validation = None
        seller_name = invoice.get("seller", "")
        if INTELLIGENCE_AVAILABLE:
            print("    ğŸ§  Intelligence: Pre-classifying from own knowledge...")
            for item in items[:3]:
                if not isinstance(item, dict):
                    continue
                # Use extracted product description, not raw text blob
                desc = item.get("description", "")
                if not desc or desc.startswith("=== ") or "&nbsp;" in desc:
                    continue  # Skip raw email/HTML text â€” not a real product description
                item_origin = item.get("origin_country", origin)
                pc_result = pre_classify(db, desc, item_origin, seller_name=seller_name)
                if not isinstance(pc_result, dict):
                    continue
                intelligence_results[desc[:50]] = pc_result
                if pc_result.get("context_text"):
                    intelligence_context += pc_result["context_text"] + "\n\n"

            # Validate documents present in extracted text
            has_fta = any(
                r.get("fta", {}).get("eligible", False)
                for r in intelligence_results.values()
                if isinstance(r, dict)
            )
            doc_validation = validate_documents(doc_text, direction="import", has_fta=has_fta)
            if doc_validation.get("missing"):
                missing_names = ", ".join(m["name_he"] for m in doc_validation["missing"])
                print(f"    ğŸ§  Intelligence: Missing documents â€” {missing_names}")

        # Get context (existing librarian search) â€” only real product descriptions
        search_terms = [
            i.get("description", "")[:50]
            for i in items[:5]
            if isinstance(i, dict) and _is_product_description(i.get("description", ""))
        ]
        tariff = query_tariff(db, search_terms)
        ministry = query_ministry_index(db)
        rules = query_classification_rules(db)

        # Enhanced knowledge search â€” skip raw text blobs
        knowledge_context = ""
        for item in items[:3]:
            if not isinstance(item, dict):
                continue
            desc = item.get("description", "")
            if desc and _is_product_description(desc):
                knowledge = full_knowledge_search(db, desc)
                knowledge_context += build_classification_context(knowledge) + "\n"

        # Merge intelligence context with librarian context
        combined_context = intelligence_context + knowledge_context

        # â”€â”€ ELIMINATION ENGINE: Walk tariff tree, narrow candidates before Agent 2 â”€â”€
        elimination_results = {}
        if ELIMINATION_ENABLED and ELIMINATION_AVAILABLE and intelligence_results:
            try:
                print("    ğŸ¯ Elimination Engine: Walking tariff tree...")
                for desc_key, pc_result in intelligence_results.items():
                    candidates = candidates_from_pre_classify(pc_result)
                    if len(candidates) < 2:
                        continue  # Need >=2 candidates to eliminate
                    # Find the matching item for product info
                    item_match = next(
                        (i for i in items if isinstance(i, dict)
                         and i.get("description", "")[:50] == desc_key),
                        items[0] if items else {},
                    )
                    if not isinstance(item_match, dict):
                        item_match = {}
                    product_info = make_product_info(item_match)
                    elim_result = eliminate(
                        db, product_info, candidates,
                        api_key=api_key, gemini_key=gemini_key,
                    )
                    elimination_results[desc_key] = elim_result
                    surv = elim_result.get("survivor_count", 0)
                    total = elim_result.get("input_count", 0)
                    print(f"       {desc_key}: {total} â†’ {surv} survivors")

                # Enrich Agent 2 context with elimination findings
                if elimination_results:
                    combined_context += _build_elimination_context(elimination_results)
                    # Session 53: Re-inject legal context with chapter-specific highlighting
                    if CUSTOMS_LAW_AVAILABLE:
                        try:
                            _elim_chapters = set()
                            for _er in elimination_results.values():
                                for _s in (_er.get("survivors") or []):
                                    _ch = _s.get("hs_code", "")[:2]
                                    if _ch.isdigit() and 1 <= int(_ch) <= 97:
                                        _elim_chapters.add(int(_ch))
                            if _elim_chapters:
                                print(f"    ğŸ“š Legal context: highlighting chapters {sorted(_elim_chapters)}")
                                combined_context += "\n\n" + format_legal_context_for_prompt(chapters=sorted(_elim_chapters))
                        except Exception:
                            pass
            except Exception as elim_err:
                print(f"    âš ï¸ Elimination engine error: {elim_err}")

        # Agent 2: Classify (Claude Sonnet 4.5 â€” core task, best quality)
        items_payload = json.dumps(items, ensure_ascii=False)
        print(f"    ğŸ·ï¸ Agent 2: Classifying {len(items)} items ({len(items_payload)} chars payload)... [Claude Sonnet 4.5]")
        if len(items) == 1:
            desc_preview = items[0].get('description', '')[:150] if isinstance(items[0], dict) else str(items[0])[:150]
            print(f"    âš ï¸ Agent 2 WARNING: Only 1 item! desc: {desc_preview}")
        classification = run_classification_agent(api_key, items, tariff, rules, combined_context, gemini_key=gemini_key)
        if not isinstance(classification, dict):
            classification = {"classifications": []}

        # Session 11: Validate HS codes against tariff database
        # First: reject non-HS-code text (e.g., Hebrew "×œ× × ×™×ª×Ÿ ×œ×¡×•×•×’")
        raw_classifications = classification.get("classifications") or []
        for c in raw_classifications:
            hs = c.get("hs_code", "")
            if hs and not _is_valid_hs(hs):
                print(f"    âš ï¸ Agent 2 returned non-HS text: '{hs}' â†’ marking unclassified")
                c["original_hs_code"] = hs
                c["hs_code"] = ""
                c["confidence"] = "× ××•×›×”"
                c["hs_warning"] = f"âš ï¸ '{hs}' ××™× ×• ×§×•×“ HS ×ª×§×™×Ÿ"

        print("    âœ… Validating HS codes against tariff database...")
        validated_classifications = validate_and_correct_classifications(db, raw_classifications)
        classification["classifications"] = validated_classifications

        # Log validation results
        for c in validated_classifications:
            if c.get('hs_corrected'):
                print(f"    âš ï¸ HS corrected: {c.get('original_hs_code')} â†’ {c.get('hs_code')}")
            elif c.get('hs_warning'):
                print(f"    âš ï¸ {c.get('hs_warning')}")

        # â”€â”€ KNOWN_FAILURES pattern guard â”€â”€
        kf_count = _check_known_failure_patterns(validated_classifications)
        if kf_count:
            print(f"    ğŸš¨ KNOWN FAILURE GUARD: {kf_count} classification(s) match known failure patterns!")

        # â”€â”€ FREE IMPORT ORDER: Verify HS codes against official API â”€â”€
        # Only for valid HS codes â€” skip text like "×œ× × ×™×ª×Ÿ ×œ×¡×•×•×’"
        free_import_results = {}
        if INTELLIGENCE_AVAILABLE:
            for c in validated_classifications[:3]:
                hs = c.get("hs_code", "")
                if hs and _is_valid_hs(hs):
                    try:
                        fio = query_free_import_order(db, hs)
                        if fio.get("found"):
                            free_import_results[hs] = fio
                    except Exception as e:
                        print(f"    âš ï¸ Free Import Order query error: {e}")

        # â”€â”€ MINISTRY ROUTING: Combine all sources into actionable guidance â”€â”€
        # Only for valid HS codes
        ministry_routing = {}
        if INTELLIGENCE_AVAILABLE:
            for c in validated_classifications[:3]:
                hs = c.get("hs_code", "")
                if hs and _is_valid_hs(hs):
                    try:
                        fio_for_hs = free_import_results.get(hs)
                        routing = route_to_ministries(db, hs, fio_for_hs)
                        if routing.get("ministries"):
                            ministry_routing[hs] = routing
                    except Exception as e:
                        print(f"    âš ï¸ Ministry routing error: {e}")

        # â”€â”€ VERIFICATION LOOP: Verify, enrich, cache every classification â”€â”€
        if VERIFICATION_AVAILABLE:
            try:
                validated_classifications = verify_all_classifications(
                    db, validated_classifications, free_import_results=free_import_results
                )
                classification["classifications"] = validated_classifications
                # Learn from verified items
                for c in validated_classifications:
                    if c.get("verification_status") in ("official", "verified"):
                        learn_from_verification(db, c)
            except Exception as e:
                print(f"    âš ï¸ Verification loop error: {e}")

        # -- BLOCK E: Verification Engine (Phase 4 + 5 + Flagging) --
        ve_results = {}
        if VERIFICATION_ENGINE_ENABLED and VERIFICATION_ENGINE_AVAILABLE:
            try:
                ve_results = run_verification_engine(
                    db, validated_classifications,
                    elimination_results=elimination_results,
                    free_import_results=free_import_results,
                    api_key=api_key, gemini_key=gemini_key,
                )
                # Apply confidence adjustments from Phase 5
                for c in validated_classifications:
                    hs = c.get("hs_code", "")
                    ve = ve_results.get(hs, {})
                    adj = ve.get("phase5", {}).get("confidence_adjustment", 0)
                    _apply_confidence_adjustment(c, adj)
            except Exception as ve_err:
                print(f"    Verification engine error (non-fatal): {ve_err}")

        # â”€â”€ LINK: Map invoice line items â†’ classifications â”€â”€
        validated_classifications = _link_invoice_to_classifications(items, validated_classifications)
        classification["classifications"] = validated_classifications

        # â”€â”€ SMART QUESTIONS: Detect ambiguity and generate elimination questions â”€â”€
        smart_questions = []
        ambiguity_info = {}
        if SMART_QUESTIONS_AVAILABLE:
            try:
                ask, ambiguity_info = should_ask_questions(
                    validated_classifications,
                    intelligence_results=intelligence_results,
                    free_import_results=free_import_results,
                    ministry_routing=ministry_routing,
                )
                if ask:
                    smart_questions = generate_smart_questions(
                        ambiguity_info,
                        validated_classifications,
                        invoice_data=invoice,
                        free_import_results=free_import_results,
                        ministry_routing=ministry_routing,
                        parsed_documents=parsed_documents,
                    )
                    print(f"    â“ Smart questions: {len(smart_questions)} questions generated "
                          f"(reason: {ambiguity_info.get('reason', 'unknown')})")
                else:
                    print(f"    âœ… Classification clear â€” no questions needed")
            except Exception as e:
                print(f"    âš ï¸ Smart questions error: {e}")

        # Agent 3: Regulatory (Gemini Flash)
        print("    âš–ï¸ Agent 3: Regulatory... [Gemini Flash]")
        regulatory = run_regulatory_agent(api_key, classification.get("classifications", []), ministry, gemini_key=gemini_key)
        if not isinstance(regulatory, dict):
            regulatory = {"regulatory": []}

        # Agent 4: FTA (Gemini Flash)
        print("    ğŸŒ Agent 4: FTA... [Gemini Flash]")
        fta = run_fta_agent(api_key, classification.get("classifications", []), origin, gemini_key=gemini_key)
        if not isinstance(fta, dict):
            fta = {"fta": []}

        # Agent 5: Risk (Gemini Flash)
        print("    ğŸš¨ Agent 5: Risk... [Gemini Flash]")
        risk = run_risk_agent(api_key, invoice, classification.get("classifications", []), gemini_key=gemini_key)
        if not isinstance(risk, dict):
            risk = {"risk": {"level": "× ××•×š", "items": []}}

        # Agent 6: Synthesis (Gemini Pro)
        print("    ğŸ“ Agent 6: Synthesis... [Gemini Pro]")
        all_results = {"invoice": invoice, "classification": classification, "regulatory": regulatory, "fta": fta, "risk": risk}

        # Include intelligence results for synthesis context
        if intelligence_results:
            all_results["intelligence"] = {
                "pre_classify": {
                    k: {
                        "top_candidate": v["candidates"][0] if v.get("candidates") else None,
                        "candidates_found": v.get("stats", {}).get("candidates_found", 0),
                        "fta_eligible": v.get("stats", {}).get("fta_eligible", False),
                        "fta": v.get("fta") if isinstance(v.get("fta"), dict) else None,
                    }
                    for k, v in intelligence_results.items()
                    if isinstance(v, dict)
                }
            }
        if doc_validation:
            all_results["document_validation"] = {
                "score": doc_validation.get("score", 0),
                "present": [p["name_he"] for p in doc_validation.get("present", [])],
                "missing": [m["name_he"] for m in doc_validation.get("missing", [])],
            }

        # Include parsed document details for synthesis
        if parsed_documents:
            all_results["parsed_documents"] = [
                {
                    "filename": pd.get("filename", ""),
                    "doc_type": pd["doc_type"],
                    "type_name": pd["type_info"]["name_he"],
                    "confidence": pd["type_info"]["confidence"],
                    "completeness_score": pd["completeness"]["score"],
                    "fields_extracted": list(pd["fields"].keys()),
                    "critical_missing": [
                        m["name_he"] for m in pd["completeness"].get("missing", [])
                        if m["importance"] == "critical"
                    ],
                }
                for pd in parsed_documents
            ]

        # Include Free Import Order official results
        if free_import_results:
            all_results["free_import_order"] = {
                hs: {
                    "authorities": [a["name"] for a in fio.get("authorities", [])],
                    "requirements": [
                        req["name"]
                        for item in fio.get("items", [])
                        for req in item.get("legal_requirements", [])
                    ][:10],
                    "decree": fio.get("decree_version", ""),
                }
                for hs, fio in free_import_results.items()
            }

        # Include ministry routing with procedures and documents
        if ministry_routing:
            all_results["ministry_routing"] = {
                hs: {
                    "risk_level": r.get("risk_level", "low"),
                    "summary_he": r.get("summary_he", ""),
                    "ministries": [
                        {
                            "name_he": m["name_he"],
                            "url": m.get("url", ""),
                            "documents_he": m.get("documents_he", []),
                            "procedure": m.get("procedure", ""),
                            "official": m.get("official", False),
                        }
                        for m in r.get("ministries", [])
                    ],
                }
                for hs, r in ministry_routing.items()
            }

        # Include smart questions for synthesis awareness
        if smart_questions:
            all_results["smart_questions"] = {
                "reason": ambiguity_info.get("reason", ""),
                "questions_count": len(smart_questions),
                "first_question": smart_questions[0]["question_he"] if smart_questions else "",
            }

        # â”€â”€ QUALITY GATE: Audit + retry before synthesis â”€â”€
        pre_send = {"agents": all_results, "invoice_data": invoice}
        audit = audit_before_send(
            pre_send, api_key=api_key, items=items,
            tariff=tariff, rules=rules, context=combined_context,
            gemini_key=gemini_key, db=db, openai_key=openai_key,
        )
        all_results["classification"]["classifications"] = audit["classifications"]

        synthesis = run_synthesis_agent(api_key, all_results, gemini_key=gemini_key)

        # Session 14: Clean synthesis text (fix typos, VAT rate, RTL spacing)
        if LANGUAGE_TOOLS_AVAILABLE:
            try:
                synthesis = _lang_checker.fix_all(synthesis)
            except Exception as e:
                print(f"    âš ï¸ Language fix_all failed: {e}")

        result = {
            "success": True,
            "agents": all_results,
            "synthesis": synthesis,
            "invoice_data": invoice,  # Session 10: Pass invoice data for validation
            "intelligence": intelligence_results,  # Session 18: Pre-classify results
            "document_validation": doc_validation,  # Session 18: Document check
            "free_import_order": free_import_results,  # Session 18: Official API results
            "ministry_routing": ministry_routing,  # Session 18: Phase C ministry routing
            "parsed_documents": parsed_documents,  # Session 19: Per-document parsing
            "smart_questions": smart_questions,  # Phase E: Elimination-based questions
            "ambiguity": ambiguity_info,  # Phase E: Ambiguity analysis
            "tracker": tracker_info,  # Shipment phase tracker
            "audit": audit,  # Quality gate results
            "elimination": elimination_results,  # Session 33 D9: Tariff tree elimination
            "verification_engine": ve_results,  # Session 34 Block E: Phase 4+5+Flagging
        }
        return result
    except Exception as e:
        print(f"    âŒ Error: {e}")
        return {"success": False, "error": str(e)}


# =============================================================================
# QUALITY GATE: Audit output before sending to customer
# =============================================================================

_CONFIDENCE_SCORES = {"×’×‘×•×”×”": 80, "×‘×™× ×•× ×™×ª": 50, "× ××•×›×”": 20}
_CONFIDENCE_TO_FLOAT = {"×’×‘×•×”×”": 0.80, "×‘×™× ×•× ×™×ª": 0.50, "× ××•×›×”": 0.20}
_HEBREW_HS_PATTERNS = re.compile(r'[\u0590-\u05FF]')


def _apply_confidence_adjustment(classification: dict, adjustment: float) -> None:
    """Apply a numeric confidence adjustment to a classification dict.

    Handles both Hebrew string confidence ('×’×‘×•×”×”'/'×‘×™× ×•× ×™×ª'/'× ××•×›×”')
    and numeric (int/float) confidence values.
    """
    if not adjustment:
        return
    conf = classification.get("confidence", "×‘×™× ×•× ×™×ª")
    if isinstance(conf, (int, float)):
        numeric = conf
    elif isinstance(conf, str):
        numeric = _CONFIDENCE_TO_FLOAT.get(conf, 0.50)
    else:
        return
    numeric = max(0.0, min(1.0, numeric + adjustment))
    # Convert back to Hebrew string for downstream email rendering
    if numeric >= 0.65:
        classification["confidence"] = "×’×‘×•×”×”"
    elif numeric >= 0.35:
        classification["confidence"] = "×‘×™× ×•× ×™×ª"
    else:
        classification["confidence"] = "× ××•×›×”"


def _retry_classification(api_key, items, tariff, rules, context, gemini_key=None):
    """Retry Agent 2 with explicit instruction to return numeric HS codes."""
    retry_instruction = """×—×©×•×‘ ×××•×“: ××ª×” ×—×™×™×‘ ×œ×”×—×–×™×¨ ×§×•×“ HS ××¡×¤×¨×™ ×‘×œ×‘×“ (6-10 ×¡×¤×¨×•×ª).
×× ××™× ×š ×‘×˜×•×—, ×”×—×–×¨ ××ª ×”×§×•×“ ×”×§×¨×•×‘ ×‘×™×•×ª×¨ ×¢× confidence "× ××•×›×”".
×œ×¢×•×œ× ××œ ×ª×—×–×™×¨ ×˜×§×¡×˜ ×›××• "×œ× × ×™×ª×Ÿ ×œ×¡×•×•×’" ×‘×©×“×” hs_code.
×× ×‘×××ª ××™×Ÿ ××¡×¤×™×§ ××™×“×¢ ×œ×¡×™×•×•×’, ×”×—×–×¨ hs_code ×¨×™×§ ""."""
    enhanced_context = retry_instruction + "\n\n" + (context or "")
    return run_classification_agent(api_key, items, tariff, rules, enhanced_context, gemini_key=gemini_key)


def audit_before_send(results, api_key=None, items=None, tariff=None, rules=None,
                      context=None, gemini_key=None, db=None, openai_key=None):
    """Quality gate â€” validate classification output before sending email.

    Returns dict:
        action: "send" | "send_with_warning" | "send_clarification" | "send_unclassified"
        classifications: cleaned/retried classification list
        warning_banner: HTML string or None
        avg_confidence: numeric 0-100
    """
    classifications = results.get("agents", {}).get("classification", {}).get("classifications", [])
    items = items or results.get("invoice_data", {}).get("items", [])

    print("    ğŸ” Quality Gate: Auditing classifications before send...")

    # â”€â”€ Step 1: Check every HS code, collect bad ones â”€â”€
    bad_indices = []
    for i, c in enumerate(classifications):
        hs = c.get("hs_code", "")
        if not hs:
            bad_indices.append(i)
            continue
        # Hebrew text in HS code field
        if _HEBREW_HS_PATTERNS.search(hs):
            print(f"    âš ï¸ QG: Hebrew text in HS code: '{hs}' (item: {c.get('item', '')[:40]})")
            c["original_hs_code"] = hs
            c["hs_code"] = ""
            c["confidence"] = "× ××•×›×”"
            c["hs_warning"] = f"âš ï¸ '{hs}' ××™× ×• ×§×•×“ HS ×ª×§×™×Ÿ"
            bad_indices.append(i)
            continue
        if not _is_valid_hs(hs):
            print(f"    âš ï¸ QG: Invalid HS format: '{hs}' (item: {c.get('item', '')[:40]})")
            c["original_hs_code"] = hs
            c["hs_code"] = ""
            c["confidence"] = "× ××•×›×”"
            c["hs_warning"] = f"âš ï¸ '{hs}' ××™× ×• ×§×•×“ HS ×ª×§×™×Ÿ"
            bad_indices.append(i)

    # â”€â”€ Step 2: Retry once if we have bad classifications and API key â”€â”€
    retried = False
    if bad_indices and api_key and tariff is not None and rules is not None:
        print(f"    ğŸ”„ QG: Retrying Agent 2 for {len(bad_indices)} unclassified items...")
        retry_result = _retry_classification(api_key, items, tariff, rules, context, gemini_key)
        if isinstance(retry_result, dict):
            retry_cls = retry_result.get("classifications", [])
            fixed = 0
            for ri, rc in enumerate(retry_cls):
                hs = rc.get("hs_code", "")
                if hs and _is_valid_hs(hs) and not _HEBREW_HS_PATTERNS.search(hs):
                    # Find matching original by item name or index
                    if ri < len(classifications):
                        orig = classifications[ri]
                        if not orig.get("hs_code") or not _is_valid_hs(orig.get("hs_code", "")):
                            orig["hs_code"] = hs
                            orig["duty_rate"] = rc.get("duty_rate", orig.get("duty_rate", ""))
                            orig["reasoning"] = rc.get("reasoning", orig.get("reasoning", ""))
                            orig["confidence"] = rc.get("confidence", "×‘×™× ×•× ×™×ª")
                            orig.pop("hs_warning", None)
                            orig["hs_retried"] = True
                            fixed += 1
            if fixed:
                print(f"    âœ… QG: Retry fixed {fixed} classifications")
                retried = True
            # Rebuild bad_indices after retry
            bad_indices = [i for i, c in enumerate(classifications)
                           if not c.get("hs_code") or not _is_valid_hs(c.get("hs_code", ""))]

    # â”€â”€ Step 3: Compute average confidence â”€â”€
    scores = []
    for c in classifications:
        conf = c.get("confidence", "×‘×™× ×•× ×™×ª")
        scores.append(_CONFIDENCE_SCORES.get(conf, 50))
    avg_confidence = sum(scores) / len(scores) if scores else 0

    # â”€â”€ Step 4: Determine action â”€â”€
    total = len(classifications)
    classified = total - len(bad_indices)

    if total == 0 or classified == 0:
        # ALL items unclassified â€” try to get candidates via GPT-4o
        print(f"    âš ï¸ QG: ALL {total} items unclassified â€” requesting candidates from GPT-4o")
        candidates_result = _get_clarification_candidates(items, context, openai_key)
        if candidates_result:
            classifications = candidates_result["classifications"]
            action = "send_clarification"
            warning_banner = candidates_result["banner"]
            print(f"    ğŸ“‹ QG: Got {len(classifications)} candidates â€” sending clarification")
        else:
            # Fallback: GPT-4o unavailable or failed
            print(f"    âš ï¸ QG: Candidate fetch failed â€” falling back to unclassified banner")
            action = "send_unclassified"
            warning_banner = _build_unclassified_banner(items)
    elif avg_confidence < 50:
        print(f"    âš ï¸ QG: Low average confidence ({avg_confidence:.0f}%) â€” adding warning")
        action = "send_with_warning"
        warning_banner = _build_low_confidence_banner(avg_confidence, classified, total)
    else:
        action = "send"
        warning_banner = None
        print(f"    âœ… QG: {classified}/{total} classified, avg confidence {avg_confidence:.0f}%")

    # Write back cleaned classifications
    results["agents"]["classification"]["classifications"] = classifications
    results["audit"] = {
        "action": action,
        "avg_confidence": round(avg_confidence),
        "classified": classified,
        "total": total,
        "retried": retried,
    }

    return {
        "action": action,
        "classifications": classifications,
        "warning_banner": warning_banner,
        "avg_confidence": avg_confidence,
    }


def _get_clarification_candidates(items, context, openai_key):
    """Ask GPT-4o for top 3 candidate HS codes when classification failed.

    Returns dict with 'classifications' list and 'banner' HTML, or None on failure.
    """
    if not openai_key:
        print("    âš ï¸ No OpenAI key â€” cannot fetch candidates")
        return None

    # Build product descriptions from items
    product_lines = []
    for i, item in enumerate(items or []):
        if isinstance(item, dict):
            desc = item.get("description", item.get("item", ""))
            if desc and _is_product_description(desc):
                product_lines.append(f"- Item {i+1}: {desc[:300]}")
    if not product_lines:
        return None

    products_text = "\n".join(product_lines)
    system_prompt = (
        "You are an Israeli customs classification expert. "
        "You classify products into Israeli HS tariff codes (×¤×¨×˜ ××›×¡). "
        "Israeli format: XX.XX.XXXXXX/X (chapter.heading.subheading/check-digit, e.g. 73.26.900002/9). "
        "Always respond in valid JSON."
    )
    user_prompt = f"""The following products could not be automatically classified.
Provide the top 3 candidate HS codes for each product.

Products:
{products_text}

{f"Additional context: {context[:2000]}" if context else ""}

Return JSON:
{{
  "candidates": [
    {{
      "item_index": 1,
      "item_description": "...",
      "options": [
        {{"hs_code": "XX.XX.XXXXXX/X", "tariff_desc_he": "×ª×™××•×¨ ××ª×¢×¨×™×£ ×”××›×¡", "reasoning": "one sentence why"}},
        {{"hs_code": "XX.XX.XXXXXX/X", "tariff_desc_he": "...", "reasoning": "..."}},
        {{"hs_code": "XX.XX.XXXXXX/X", "tariff_desc_he": "...", "reasoning": "..."}}
      ],
      "distinguishing_info": "What information would distinguish between these candidates"
    }}
  ]
}}"""

    try:
        raw = call_chatgpt(openai_key, system_prompt, user_prompt, max_tokens=2000, model="gpt-4o")
        if not raw:
            return None

        import json as _json
        # Strip markdown code fences if present
        cleaned = raw.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("\n", 1)[1] if "\n" in cleaned else cleaned[3:]
            if cleaned.endswith("```"):
                cleaned = cleaned[:-3]
            cleaned = cleaned.strip()

        data = _json.loads(cleaned)
        candidates_list = data.get("candidates", [])
        if not candidates_list:
            return None

        # Build classifications from candidates (use first option per item as primary)
        new_classifications = []
        for cand in candidates_list:
            options = cand.get("options", [])
            if options:
                top = options[0]
                new_classifications.append({
                    "item": cand.get("item_description", ""),
                    "hs_code": top.get("hs_code", "").replace(".", "").replace("/", "")[:10],
                    "tariff_description": top.get("tariff_desc_he", ""),
                    "reasoning": top.get("reasoning", ""),
                    "confidence": "× ××•×›×”",
                    "clarification_candidates": options,
                    "distinguishing_info": cand.get("distinguishing_info", ""),
                })

        if not new_classifications:
            return None

        # Build clarification banner HTML
        banner = _build_clarification_banner(candidates_list)
        return {"classifications": new_classifications, "banner": banner}

    except Exception as e:
        print(f"    âš ï¸ GPT-4o candidates error: {e}")
        return None


def _build_clarification_banner(candidates_list):
    """Build HTML banner showing candidate HS codes and clarification questions."""
    rows_html = ""
    questions_html = ""
    for cand in candidates_list:
        item_desc = cand.get("item_description", "")[:80]
        dist_info = cand.get("distinguishing_info", "")
        options = cand.get("options", [])

        for j, opt in enumerate(options):
            letter = chr(0x05D0 + j)  # ×, ×‘, ×’
            rows_html += (
                f'<tr><td style="padding:8px;border:1px solid #dee2e6">{letter}</td>'
                f'<td style="padding:8px;border:1px solid #dee2e6;font-family:monospace">{get_israeli_hs_format(opt.get("hs_code", ""))}</td>'
                f'<td style="padding:8px;border:1px solid #dee2e6">{opt.get("tariff_desc_he", "")}</td>'
                f'<td style="padding:8px;border:1px solid #dee2e6">{opt.get("reasoning", "")}</td></tr>'
            )

        if dist_info:
            questions_html += f'<li style="margin:5px 0"><strong>{item_desc}:</strong> {dist_info}</li>'

    return f'''<div dir="rtl" style="background:#fff3cd;border:2px solid #ffc107;border-radius:5px;padding:20px;margin-bottom:20px;font-family:Arial,sans-serif">
        <h3 style="color:#856404;margin:0">âš ï¸ ×“×¨×•×©×” ×”×‘×”×¨×” ×œ×¡×™×•×•×’ ××“×•×™×§</h3>
        <p style="color:#856404;margin:10px 0">×”××¢×¨×›×ª ××¦××” ××¡×¤×¨ ××¤×©×¨×•×™×•×ª ×¡×™×•×•×’. × × ×œ×‘×—×•×¨ ××• ×œ×¡×¤×§ ××™×“×¢ × ×•×¡×£:</p>
        <table dir="rtl" style="width:100%;border-collapse:collapse;margin:10px 0;background:#fff">
            <tr style="background:#f8f9fa">
                <th style="padding:8px;border:1px solid #dee2e6;width:30px">#</th>
                <th style="padding:8px;border:1px solid #dee2e6">×¤×¨×˜ ××›×¡</th>
                <th style="padding:8px;border:1px solid #dee2e6">×ª×™××•×¨ ××ª×¢×¨×™×£</th>
                <th style="padding:8px;border:1px solid #dee2e6">× ×™××•×§</th>
            </tr>
            {rows_html}
        </table>
        <div style="background:#fff;border:1px solid #ffc107;border-radius:3px;padding:15px;margin-top:10px">
            <p style="margin:0 0 5px 0"><strong>×œ×¦×•×¨×š ×¡×™×•×•×’ ××“×•×™×§, × × ×”×‘×”×™×¨×•:</strong></p>
            <ul style="margin:5px 0;padding-right:20px">{questions_html}</ul>
            <p style="margin:10px 0 0 0;color:#856404">× × ×œ×”×©×™×‘ ×œ××™×™×œ ×–×” ×¢× ×”×ª×©×•×‘×•×ª.</p>
        </div>
    </div>'''


def _build_unclassified_banner(items):
    """HTML banner for when all items are unclassified."""
    item_list = ""
    for item in (items or [])[:5]:
        if isinstance(item, dict):
            desc = item.get("description", item.get("item", ""))
            if desc and _is_product_description(desc):
                item_list += f"<li>{desc[:100]}</li>"
    items_html = f"<ul>{item_list}</ul>" if item_list else ""

    return f'''<div style="background:#f8d7da;border:2px solid #dc3545;border-radius:5px;padding:20px;margin-bottom:20px">
        <h3 style="color:#721c24;margin:0">âš ï¸ ×œ× ×”×¦×œ×—× ×• ×œ×¡×•×•×’ ××ª ×”×¤×¨×™×˜×™×</h3>
        <p style="color:#721c24;margin:10px 0 0 0">×”××¢×¨×›×ª ×œ× ×”×¦×œ×™×—×” ×œ×§×‘×•×¢ ×§×•×“×™ HS ×¢×‘×•×¨ ×”×¤×¨×™×˜×™× ×©×œ×”×œ×Ÿ.
        ×™×© ×œ×¡×¤×§ ×ª×™××•×¨ ××¤×•×¨×˜ ×™×•×ª×¨ ×©×œ ×”××•×¦×¨×™×, ×›×•×œ×œ: ×—×•××¨ ×’×œ×, ×©×™××•×© ××™×•×¢×“, ×”×¨×›×‘.</p>
        {items_html}
        <p style="color:#721c24;margin:10px 0 0 0"><strong>×× × ×”×©×™×‘×• ×¢× ××™×“×¢ × ×•×¡×£ ×•× ×—×–×•×¨ ×¢× ×¡×™×•×•×’.</strong></p>
    </div>'''


def _build_low_confidence_banner(avg_confidence, classified, total):
    """HTML warning banner for low average confidence."""
    return f'''<div style="background:#fff3cd;border:2px solid #ffc107;border-radius:5px;padding:15px;margin-bottom:20px">
        <h3 style="color:#856404;margin:0">âš ï¸ ×¨××ª ×•×“××•×ª × ××•×›×” ({avg_confidence:.0f}%)</h3>
        <p style="color:#856404;margin:10px 0 0 0">×¡×•×•×’×• {classified} ××ª×•×š {total} ×¤×¨×™×˜×™×, ××š ×¨××ª ×”×•×•×“××•×ª ×”×××•×¦×¢×ª × ××•×›×”.
        ××•××œ×¥ ×œ×‘×“×•×§ ××ª ×”×¡×™×•×•×’×™× ×œ×¤× ×™ ×©×™××•×©. ×™×™×ª×›×Ÿ ×©× ×“×¨×© ××™×“×¢ × ×•×¡×£ ×¢×œ ×”××•×¦×¨×™×.</p>
    </div>'''


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SESSION 40 â€” Block H: Classification Result Email (Redesigned)
# Table-based, Outlook-safe, RTL, matches tracker_email.py design language.
# Section builders return <tr>â€¦</tr> blocks; orchestrator composes them.
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•


def _cls_html_open():
    """Opening HTML/body/table wrapper (Outlook-safe, RTL) â€” local fallback."""
    return ('<!DOCTYPE html>\n'
            '<html dir="rtl" lang="he">\n'
            '<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>\n'
            '<body dir="rtl" style="margin:0;padding:0;background:#f5f5f5;font-family:Arial,Helvetica,sans-serif;">\n'
            '<table width="100%" cellpadding="0" cellspacing="0" style="background:#f5f5f5;padding:20px 0;">\n'
            '<tr><td align="center">\n'
            '<table width="640" cellpadding="0" cellspacing="0" style="max-width:640px;width:100%;'
            'background:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.1);">\n')


def _cls_html_close():
    """Closing tags matching _cls_html_open."""
    return '</table>\n</td></tr></table>\n</body></html>'


def _cls_header(tracking_code):
    """Section 1: Branded header with RPA-PORT logo and tracking code."""
    return f"""
<!-- HEADER -->
<tr><td style="background:{_RPA_BLUE};padding:0;">
  <table width="100%" cellpadding="0" cellspacing="0">
  <tr>
    <td style="padding:20px 30px;" valign="middle">
      <span style="display:inline-block;vertical-align:middle;width:48px;height:48px;line-height:48px;text-align:center;background:#1e3a5f;color:#ffffff;font-size:20px;font-weight:bold;border-radius:6px;font-family:Georgia,'Times New Roman',serif;">RCB</span>
      <span style="display:inline-block;vertical-align:middle;padding-left:12px;">
        <span style="color:#ffffff;font-size:18px;font-weight:bold;display:block;">R.P.A. PORT LTD</span>
        <span style="color:#aed6f1;font-size:13px;display:block;">&#1491;&#1493;&#1524;&#1495; &#1505;&#1497;&#1493;&#1493;&#1490; &#1502;&#1499;&#1505; &mdash; RCB</span>
      </span>
    </td>
    <td align="left" style="padding:20px 30px;" valign="middle">
      <table cellpadding="0" cellspacing="0" style="background:rgba(255,255,255,0.15);border-radius:8px;">
      <tr><td style="padding:6px 14px;text-align:center;">
        <span style="font-size:10px;color:#aed6f1;display:block;font-family:'Courier New',monospace;">TRACKING</span>
        <span style="font-size:13px;color:#ffffff;font-family:'Courier New',monospace;font-weight:bold;letter-spacing:1px;">{tracking_code or ''}</span>
      </td></tr>
      </table>
    </td>
  </tr>
  </table>
</td></tr>
"""


def _cls_shipment_info(invoice_data):
    """Shipment info: direction, freight, parties, reference number."""
    if not invoice_data:
        return ""
    direction = invoice_data.get('direction', 'unknown')
    dir_he = "\u05d9\u05d1\u05d5\u05d0 \u05dc\u05d9\u05e9\u05e8\u05d0\u05dc" if direction == 'import' else "\u05d9\u05e6\u05d5\u05d0 \u05de\u05d9\u05e9\u05e8\u05d0\u05dc" if direction == 'export' else "\u05dc\u05d0 \u05d9\u05d3\u05d5\u05e2"
    freight = invoice_data.get('freight_type', 'unknown')
    freight_he = "\u05d9\u05dd" if freight == 'sea' else "\u05d0\u05d5\u05d5\u05d9\u05e8" if freight == 'air' else "\u05dc\u05d0 \u05d9\u05d3\u05d5\u05e2"
    seller = (invoice_data.get('seller', '') or '\u05dc\u05d0 \u05d9\u05d3\u05d5\u05e2')[:35]
    buyer = (invoice_data.get('buyer', '') or '\u05dc\u05d0 \u05d9\u05d3\u05d5\u05e2')[:35]
    bl = invoice_data.get('bl_number', '')
    awb = invoice_data.get('awb_number', '')
    ref = f"B/L: {bl}" if bl else f"AWB: {awb}" if awb else ""

    def _kv(label, value):
        v = value if value else '&#8212;'
        return (f'  <tr>'
                f'<td style="padding:6px 12px;color:{_RPA_BLUE};font-weight:bold;font-size:13px;'
                f'width:35%;border-bottom:1px solid #f0f0f0;">{label}</td>'
                f'<td style="padding:6px 12px;color:#333;font-size:13px;'
                f'border-bottom:1px solid #f0f0f0;">{v}</td>'
                f'</tr>\n')

    html = f"""
<!-- SHIPMENT INFO -->
<tr><td style="padding:20px 30px 10px;">
  <table width="100%" cellpadding="0" cellspacing="0"
         style="font-size:14px;font-weight:bold;color:{_RPA_BLUE};margin-bottom:8px;">
  <tr><td>\u05e4\u05e8\u05d8\u05d9 \u05de\u05e9\u05dc\u05d5\u05d7</td></tr>
  </table>
  <table width="100%" cellpadding="0" cellspacing="0"
         style="border:1px solid #e0e0e0;border-radius:4px;overflow:hidden;">
"""
    html += _kv("\u05db\u05d9\u05d5\u05d5\u05df", dir_he)
    html += _kv("\u05d4\u05d5\u05d1\u05dc\u05d4", freight_he)
    html += _kv("\u05de\u05d5\u05db\u05e8", seller)
    html += _kv("\u05e7\u05d5\u05e0\u05d4", buyer)
    if ref:
        html += _kv("\u05de\u05e1\u05f3 \u05de\u05e9\u05dc\u05d5\u05d7", f'<span style="font-family:\'Courier New\',monospace;">{ref}</span>')
    html += """  </table>
</td></tr>"""
    return html


def _cls_invoice_validation(invoice_validation):
    """Invoice validation score bar."""
    if not invoice_validation:
        return ""
    score = invoice_validation.get('score', 0)
    is_valid = invoice_validation.get('is_valid', False)
    missing = invoice_validation.get('missing_fields', [])
    bar_color = _COLOR_OK if is_valid else _COLOR_WARN
    text_color = "#166534" if is_valid else "#92400e"
    label = "\u05d7\u05e9\u05d1\u05d5\u05df \u05ea\u05e7\u05d9\u05df" if is_valid else "\u05d7\u05e9\u05d1\u05d5\u05df \u05d7\u05dc\u05e7\u05d9"
    icon = "&#10003;" if is_valid else "&#9888;"

    html = f"""
<!-- INVOICE VALIDATION -->
<tr><td style="padding:10px 30px;">
  <table width="100%" cellpadding="0" cellspacing="0"
         style="border:1px solid {bar_color};border-radius:4px;overflow:hidden;">
  <tr>
    <td style="padding:12px 16px;vertical-align:middle;">
      <span style="font-size:15px;font-weight:bold;color:{text_color};">{icon} {label}</span>
    </td>
    <td style="padding:12px 16px;width:160px;" align="left">
      <table width="100%" cellpadding="0" cellspacing="0">
      <tr><td style="font-size:11px;color:{text_color};padding-bottom:4px;">\u05e6\u05d9\u05d5\u05df: {score}/100</td></tr>
      <tr><td>
        <table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f0f0;border-radius:10px;overflow:hidden;">
        <tr><td style="width:{score}%;background:{bar_color};height:8px;border-radius:10px;">&nbsp;</td>
            <td style="height:8px;">&nbsp;</td></tr>
        </table>
      </td></tr>
      </table>
    </td>
  </tr>"""
    if not is_valid and missing:
        html += f"""  <tr><td colspan="2" style="padding:8px 16px 12px;border-top:1px solid #f0f0f0;">
      <span style="font-size:12px;color:{text_color};">\u05d7\u05e1\u05e8\u05d9\u05dd {len(missing)} \u05e9\u05d3\u05d5\u05ea:</span><br>"""
        for field in missing[:5]:
            html += f'      <span style="display:inline-block;background:#fff;border:1px solid {bar_color};border-radius:12px;padding:2px 10px;margin:3px;font-size:11px;color:{text_color};">{field}</span>\n'
        html += "    </td></tr>"
    html += """  </table>
</td></tr>"""
    return html


def _cls_synthesis(synthesis):
    """Synthesis section â€” blue-bordered summary text."""
    if not synthesis:
        return ""
    if LANGUAGE_TOOLS_AVAILABLE:
        try:
            synthesis = _lang_checker.fix_all(synthesis)
        except Exception:
            pass
    return f"""
<!-- SYNTHESIS -->
<tr><td style="padding:10px 30px;">
  <table width="100%" cellpadding="0" cellspacing="0"
         style="border:1px solid #d4e3f5;border-radius:4px;overflow:hidden;">
  <tr>
    <td style="width:4px;background:{_RPA_BLUE};"></td>
    <td style="padding:14px 18px;font-size:14px;line-height:1.7;color:#333;background:#f8faff;">
      {synthesis}
    </td>
  </tr>
  </table>
</td></tr>"""


def _cls_section_title(title, color=None):
    """Render a section title with gradient underline."""
    c = color or _RPA_BLUE
    accent = _RPA_ACCENT if c == _RPA_BLUE else c
    return f"""
<tr><td style="padding:20px 30px 5px;">
  <table width="100%" cellpadding="0" cellspacing="0">
  <tr><td style="font-size:16px;font-weight:bold;color:{c};">{title}</td></tr>
  <tr><td style="padding-top:4px;">
    <table width="50" cellpadding="0" cellspacing="0">
    <tr><td style="height:3px;background:{accent};border-radius:2px;"></td></tr>
    </table>
  </td></tr>
  </table>
</td></tr>"""


def _cls_result_table(card_items, using_enriched):
    """Section 2: Classification result table â€” per-item cards with HS code + confidence."""
    if not card_items:
        return ""

    html = _cls_section_title("\u05e1\u05d9\u05d5\u05d5\u05d2\u05d9\u05dd")

    for c in card_items:
        conf = c.get("confidence", "\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9\u05ea")
        if conf == "\u05d2\u05d1\u05d5\u05d4\u05d4":
            conf_color = _COLOR_OK
            conf_width = "100"
        elif conf == "\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9\u05ea":
            conf_color = _COLOR_WARN
            conf_width = "66"
        else:
            conf_color = _COLOR_ERR
            conf_width = "33"

        hs_display = get_israeli_hs_format(c.get("hs_code", ""))
        line_num = c.get("line_number", 0)
        item_name = (c.get("item", "") or c.get("description", ""))[:55]
        tariff_text = c.get("tariff_text_he", "") or c.get("official_description_he", "")
        qty = c.get("quantity", "")
        unit_px = c.get("unit_price", "")
        item_origin = c.get("origin_country", "")
        qty_price = f'{qty} &times; ${unit_px}' if qty and unit_px else (str(qty) if qty else '')

        # Verification status badge
        v_status = c.get('verification_status', '')
        hs_badge = ""
        if c.get('hs_corrected'):
            hs_badge = '<span style="display:inline-block;background:#fef3c7;color:#92400e;font-size:10px;padding:2px 8px;border-radius:10px;">&#9888; \u05ea\u05d5\u05e7\u05df</span>'
        elif v_status in ('official', 'verified'):
            hs_badge = '<span style="display:inline-block;background:#dcfce7;color:#166534;font-size:10px;padding:2px 8px;border-radius:10px;">&#10003; \u05d0\u05d5\u05de\u05ea</span>'
        elif c.get('hs_warning'):
            hs_badge = '<span style="display:inline-block;background:#fee2e2;color:#991b1b;font-size:10px;padding:2px 8px;border-radius:10px;">&#9888; \u05dc\u05d0 \u05d0\u05d5\u05de\u05ea</span>'
        elif c.get('hs_validated') and c.get('hs_exact_match'):
            hs_badge = '<span style="display:inline-block;background:#dcfce7;color:#166534;font-size:10px;padding:2px 8px;border-radius:10px;">&#10003; \u05d0\u05d5\u05de\u05ea</span>'

        # Purchase tax
        pt = c.get("purchase_tax", {})
        if isinstance(pt, dict) and pt.get("applies"):
            pt_display = pt.get("rate_he", "\u05d7\u05dc")
            pt_note = pt.get("note_he", "")
            if pt_note:
                pt_display += f'<br><span style="color:#888;font-size:10px;">{pt_note[:25]}</span>'
        else:
            pt_display = '<span style="color:#aaa;">\u05dc\u05d0 \u05d7\u05dc</span>'

        vat_display = c.get("vat_rate", "") or "\u2014"

        # Seller/Buyer row
        seller_buyer = ""
        if using_enriched:
            _s = c.get("seller", "")
            _b = c.get("buyer", "")
            if _s or _b:
                parts = []
                if _s:
                    parts.append(f'\u05de\u05d5\u05db\u05e8: {_s}')
                if _b:
                    parts.append(f'\u05e7\u05d5\u05e0\u05d4: {_b}')
                seller_buyer = f'<div style="font-size:11px;color:#555;margin-top:3px;">{" &nbsp;|&nbsp; ".join(parts)}</div>'

        # Per-item card
        html += f"""
<tr><td style="padding:4px 30px;">
  <table width="100%" cellpadding="0" cellspacing="0"
         style="border:1px solid #e0e0e0;border-radius:4px;overflow:hidden;">
  <!-- Item header -->
  <tr style="background:#f8faff;">
    <td colspan="3" style="padding:12px 16px;border-bottom:1px solid #e5e7eb;">
      <table width="100%" cellpadding="0" cellspacing="0">
      <tr>
        <td style="vertical-align:middle;">"""
        if line_num:
            html += f'          <span style="display:inline-block;background:{_RPA_BLUE};color:#fff;font-size:11px;font-weight:bold;padding:2px 10px;border-radius:10px;margin-left:8px;">\u05e9\u05d5\u05e8\u05d4 {line_num}</span>'
        html += f"""
          <span style="font-size:14px;font-weight:bold;color:#0f2439;">{item_name}</span>
        </td>"""
        if qty_price:
            html += f"""
        <td style="text-align:left;vertical-align:middle;">
          <span style="font-size:12px;color:#555;font-family:'Courier New',monospace;">{qty_price}</span>
        </td>"""
        html += """
      </tr>
      </table>"""
        inv_desc = c.get("invoice_description", "")
        if inv_desc and inv_desc.lower().strip() != (c.get("item", "") or "").lower().strip():
            html += f'      <div style="font-size:11px;color:#888;margin-top:4px;">\u05d7\u05e9\u05d1\u05d5\u05e0\u05d9\u05ea: {inv_desc[:80]}</div>'
        if item_origin:
            html += f'      <div style="font-size:11px;color:#888;margin-top:3px;">\u05de\u05e7\u05d5\u05e8: {item_origin}</div>'
        html += seller_buyer
        html += f"""
    </td>
  </tr>
  <!-- HS Code + Taxes + Confidence -->
  <tr>
    <td style="padding:14px 16px;width:40%;vertical-align:top;">
      <span style="font-size:10px;color:#888;text-transform:uppercase;letter-spacing:0.5px;">\u05e7\u05d5\u05d3 HS</span><br>
      <span style="font-family:'Courier New',monospace;font-size:18px;font-weight:bold;color:{_RPA_BLUE};letter-spacing:0.5px;">{hs_display}</span>
      {hs_badge}"""
        if tariff_text:
            html += f"""
      <div style="font-size:11px;color:#555;margin-top:4px;line-height:1.4;font-style:italic;">{tariff_text[:120]}</div>"""
        html += f"""
    </td>
    <td style="padding:14px 8px;vertical-align:top;border-right:1px solid #f0f0f0;border-left:1px solid #f0f0f0;">
      <table width="100%" cellpadding="0" cellspacing="0">
      <tr><td style="padding:2px 0;"><span style="font-size:10px;color:#888;">\u05de\u05db\u05e1</span><br><strong style="font-size:14px;color:#0f2439;">{c.get("duty_rate", "")}</strong></td></tr>
      <tr><td style="padding:2px 0;border-top:1px solid #f0f0f0;"><span style="font-size:10px;color:#888;">\u05de\u05e1 \u05e7\u05e0\u05d9\u05d9\u05d4</span><br><span style="font-size:13px;color:#333;">{pt_display}</span></td></tr>
      <tr><td style="padding:2px 0;border-top:1px solid #f0f0f0;"><span style="font-size:10px;color:#888;">\u05de\u05e2\u05f4\u05de</span><br><strong style="font-size:14px;color:#0f2439;">{vat_display}</strong></td></tr>
      </table>
    </td>
    <td style="padding:14px 16px;width:25%;vertical-align:top;">
      <span style="font-size:10px;color:#888;">\u05d5\u05d3\u05d0\u05d5\u05ea</span><br>
      <span style="font-size:14px;font-weight:bold;color:{conf_color};">{conf}</span>
      <table width="100%" cellpadding="0" cellspacing="0" style="margin-top:6px;">
      <tr><td>
        <table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f0f0;border-radius:10px;overflow:hidden;">
        <tr><td style="width:{conf_width}%;background:{conf_color};height:8px;border-radius:10px;">&nbsp;</td>
            <td style="height:8px;">&nbsp;</td></tr>
        </table>
      </td></tr>
      </table>
    </td>
  </tr>
  </table>
</td></tr>"""

    return html


def _cls_regulatory(ministry_routing, regulatory):
    """Section 3: Regulatory flags table."""
    has_routing = bool(ministry_routing)
    if not has_routing and not regulatory:
        return ""

    html = _cls_section_title("\u05e8\u05d2\u05d5\u05dc\u05e6\u05d9\u05d4 \u05d5\u05d0\u05d9\u05e9\u05d5\u05e8\u05d9\u05dd")

    if has_routing:
        for hs_code, routing in ministry_routing.items():
            ministries = routing.get("ministries", [])
            if not ministries:
                continue
            summary = routing.get("summary_he", "")
            html += f"""
<tr><td style="padding:4px 30px;">
  <table width="100%" cellpadding="0" cellspacing="0"
         style="border:1px solid #e0e0e0;border-radius:4px;overflow:hidden;">
  <tr><td style="padding:10px 16px;background:#f8faff;border-bottom:1px solid #e5e7eb;">
    <span style="font-family:'Courier New',monospace;font-weight:bold;color:{_RPA_BLUE};">{get_israeli_hs_format(hs_code)}</span>"""
            if summary:
                html += f' <span style="font-size:12px;color:#555;margin-right:12px;">{summary[:120]}</span>'
            html += "</td></tr>"
            for m in ministries:
                official = '<span style="color:#166534;font-size:10px;background:#dcfce7;padding:1px 6px;border-radius:8px;margin-left:4px;">API</span>' if m.get("official") else ''
                m_name = m.get("name_he", m.get("name", ""))
                docs = m.get("documents_he", [])
                proc = m.get("procedure", "")
                url = m.get("url", "")
                html += f'  <tr><td style="padding:8px 16px;border-bottom:1px solid #f0f0f0;">'
                html += f'<span style="font-weight:bold;color:#1e40af;font-size:13px;">{official}{m_name}</span>'
                if docs:
                    html += '<br>'
                    for d in docs[:4]:
                        html += f'<span style="display:inline-block;background:#eff6ff;border:1px solid #bfdbfe;border-radius:12px;padding:2px 10px;margin:2px;font-size:11px;color:#1e40af;">{d}</span>'
                if proc:
                    html += f'<br><span style="font-size:11px;color:#888;">{proc[:100]}</span>'
                if url:
                    html += f' <a href="{url}" style="font-size:11px;color:#2563eb;text-decoration:none;">\u05d0\u05ea\u05e8 \u05d4\u05de\u05e9\u05e8\u05d3</a>'
                html += "</td></tr>\n"
            html += "  </table>\n</td></tr>"
    elif regulatory:
        for r in regulatory:
            ministries_list = [m for m in r.get("ministries", []) if m.get("required")]
            if not ministries_list:
                continue
            html += f"""
<tr><td style="padding:4px 30px;">
  <table width="100%" cellpadding="0" cellspacing="0"
         style="border:1px solid #e0e0e0;border-radius:4px;overflow:hidden;">
  <tr><td style="padding:10px 16px;background:#f8faff;">
    <span style="font-family:'Courier New',monospace;font-weight:bold;color:{_RPA_BLUE};">{get_israeli_hs_format(r.get("hs_code", ""))}</span>
  </td></tr>"""
            for m in ministries_list:
                html += f'  <tr><td style="padding:6px 16px;border-top:1px solid #f0f0f0;"><span style="display:inline-block;background:#eff6ff;border:1px solid #bfdbfe;border-radius:12px;padding:3px 12px;font-size:12px;color:#1e40af;">{m.get("name")} &mdash; {m.get("regulation", "")}</span></td></tr>\n'
            html += "  </table>\n</td></tr>"

    return html


def _cls_cross_reference(results, card_items):
    """Section 4: Cross-reference EU TARIC / US HTS."""
    cross_ref = results.get("cross_reference")
    if not cross_ref:
        return ""
    # Check if any items have cross-ref data
    has_data = any(c.get("cross_ref_adjustment") is not None for c in card_items)
    if not has_data:
        return ""

    html = _cls_section_title("\u05d1\u05d3\u05d9\u05e7\u05d4 \u05e6\u05d5\u05dc\u05d1\u05ea (EU / US)")
    html += f"""
<tr><td style="padding:8px 30px;">
  <table width="100%" cellpadding="0" cellspacing="0"
         style="border:1px solid #e0e0e0;border-radius:4px;overflow:hidden;font-size:12px;">
  <tr style="background:{_RPA_BLUE};color:#fff;">
    <td style="padding:8px 12px;font-weight:bold;">HS Code</td>
    <td style="padding:8px 12px;font-weight:bold;text-align:center;">EU TARIC</td>
    <td style="padding:8px 12px;font-weight:bold;text-align:center;">US HTS</td>
    <td style="padding:8px 12px;font-weight:bold;text-align:center;">\u05d4\u05ea\u05d0\u05de\u05d4</td>
  </tr>"""

    for i, c in enumerate(card_items[:5]):
        hs = c.get("hs_code", "")
        if not hs:
            continue
        bg = "#f8f9fa" if i % 2 == 0 else "#ffffff"
        adj = c.get("cross_ref_adjustment")
        note = c.get("cross_ref_note", "")

        eu_key = f"eu_{hs}"
        us_key = f"us_{hs}"
        eu_found = eu_key in cross_ref
        us_found = us_key in cross_ref

        eu_icon = f'<span style="color:{_COLOR_OK};font-weight:bold;">&#10003;</span>' if eu_found else '<span style="color:#ccc;">&#8212;</span>'
        us_icon = f'<span style="color:{_COLOR_OK};font-weight:bold;">&#10003;</span>' if us_found else '<span style="color:#ccc;">&#8212;</span>'

        if adj is not None:
            if adj > 0.08:
                match_color = _COLOR_OK
                match_text = f"+{int(adj * 100)}%"
            elif adj > 0:
                match_color = _COLOR_WARN
                match_text = f"+{int(adj * 100)}%"
            else:
                match_color = _COLOR_ERR
                match_text = f"{int(adj * 100)}%"
        else:
            match_color = "#ccc"
            match_text = "&#8212;"

        html += f"""  <tr style="background:{bg};">
    <td style="padding:6px 12px;font-family:'Courier New',monospace;font-weight:bold;color:{_RPA_BLUE};">{get_israeli_hs_format(hs)}</td>
    <td style="padding:6px 12px;text-align:center;">{eu_icon}</td>
    <td style="padding:6px 12px;text-align:center;">{us_icon}</td>
    <td style="padding:6px 12px;text-align:center;color:{match_color};font-weight:bold;">{match_text}</td>
  </tr>
"""

    html += """  </table>
</td></tr>"""
    return html


def _cls_customs_value(invoice_data, results):
    """Section 5: Customs value with BOI exchange rate."""
    if not invoice_data:
        return ""
    currency = (invoice_data.get("currency", "") or "").upper()
    if not currency or currency in ("ILS", "NIS", ""):
        return ""

    # Calculate total from items
    items = invoice_data.get("items", [])
    total_value = 0
    for item in (items or []):
        if isinstance(item, dict):
            try:
                t = float(str(item.get("total", 0)).replace(",", ""))
                total_value += t
            except (ValueError, TypeError):
                pass
    if not total_value:
        try:
            total_value = float(str(invoice_data.get("total_value", 0)).replace(",", ""))
        except (ValueError, TypeError):
            pass

    # BOI rate from pre-enrichment
    pre_enrichment = results.get("pre_enrichment") or {}
    boi_data = pre_enrichment.get("boi_rate") or {}
    boi_rate = boi_data.get("rate") if boi_data else None
    boi_date = boi_data.get("lastUpdate", "") if boi_data else ""

    html = _cls_section_title("\u05e2\u05e8\u05da \u05de\u05db\u05e1")
    html += """
<tr><td style="padding:8px 30px;">
  <table width="100%" cellpadding="0" cellspacing="0"
         style="border:1px solid #e0e0e0;border-radius:4px;overflow:hidden;">"""

    def _vr(label, value, bold=False):
        s = "font-weight:bold;font-size:15px;" if bold else "font-size:13px;"
        return (f'  <tr><td style="padding:8px 16px;color:{_RPA_BLUE};font-weight:bold;font-size:13px;'
                f'width:45%;border-bottom:1px solid #f0f0f0;">{label}</td>'
                f'<td style="padding:8px 16px;color:#333;{s}'
                f'border-bottom:1px solid #f0f0f0;">{value}</td></tr>\n')

    html += _vr("\u05de\u05d8\u05d1\u05e2 \u05d7\u05e9\u05d1\u05d5\u05e0\u05d9\u05ea", currency)
    if total_value:
        html += _vr("\u05e1\u05d4\u05f4\u05db \u05d7\u05e9\u05d1\u05d5\u05e0\u05d9\u05ea", f'{total_value:,.2f} {currency}')
    if boi_rate:
        html += _vr("\u05e9\u05e2\u05e8 \u05d1\u05e0\u05e7 \u05d9\u05e9\u05e8\u05d0\u05dc", f'{boi_rate} ILS/{currency}')
        if boi_date:
            html += _vr("\u05ea\u05d0\u05e8\u05d9\u05da \u05e9\u05e2\u05e8", str(boi_date)[:16])
        if total_value:
            try:
                customs_value_ils = total_value * float(boi_rate)
                html += _vr("\u05e2\u05e8\u05da \u05de\u05db\u05e1 (ILS)", f'&#8362; {customs_value_ils:,.2f}', bold=True)
            except (ValueError, TypeError):
                pass
    else:
        html += _vr("\u05e9\u05e2\u05e8 \u05d1\u05e0\u05e7 \u05d9\u05e9\u05e8\u05d0\u05dc", '<span style="color:#999;">\u05de\u05de\u05ea\u05d9\u05df \u05dc\u05e2\u05d3\u05db\u05d5\u05df</span>')

    html += """  </table>
</td></tr>"""
    return html


def _cls_justification_details(card_items, using_enriched):
    """Section 6: Per-item justification, challenge, gaps, verification flags."""
    if not card_items:
        return ""

    # Check if any items have justification content
    has_any = False
    for c in card_items:
        j = c.get("justification")
        ch = c.get("challenge")
        if (j and j.get("chain")) or (ch and ch.get("alternatives")):
            has_any = True
            break
        if using_enriched and (c.get("ve_flags") or c.get("ve_phase4")):
            has_any = True
            break

    if not has_any:
        return ""

    html = _cls_section_title("\u05e0\u05d9\u05de\u05d5\u05e7 \u05de\u05e9\u05e4\u05d8\u05d9")

    for c in card_items:
        hs = c.get("hs_code", "")
        item_name = (c.get("item", "") or c.get("description", ""))[:50]
        item_justification = c.get("justification")
        item_challenge = c.get("challenge")

        has_content = False
        item_html = ""

        # Justification chain + challenge + gaps + cross-check
        try:
            from lib.report_builder import (
                build_justification_html, build_challenge_html,
                build_gaps_summary_html, build_cross_check_badge,
            )
            if item_justification and item_justification.get("chain"):
                item_html += build_justification_html(item_justification)
                has_content = True
            if item_challenge and item_challenge.get("alternatives"):
                item_html += build_challenge_html(item_challenge)
                has_content = True
            if item_justification and item_justification.get("gaps"):
                item_html += build_gaps_summary_html(item_justification["gaps"])
                has_content = True
            cc_badge = build_cross_check_badge(c)
            if cc_badge:
                item_html += f'<div style="margin-top:6px;">{cc_badge}</div>'
                has_content = True
        except ImportError:
            pass
        except Exception:
            pass

        # Verification flags (Block E)
        if using_enriched:
            try:
                ve_html = build_verification_flags_html(c)
                if ve_html:
                    item_html += f'<div style="margin-top:8px;">{ve_html}</div>'
                    has_content = True
            except Exception:
                pass

        # Per-item ministry approvals
        if using_enriched:
            item_ministries = c.get("ministries", [])
            if item_ministries:
                item_html += '<div style="margin-top:12px;padding-top:12px;border-top:1px solid #f0f0f0;">'
                item_html += '<span style="font-size:11px;color:#888;">\u05d0\u05d9\u05e9\u05d5\u05e8\u05d9 \u05de\u05e9\u05e8\u05d3\u05d9\u05dd \u05e0\u05d3\u05e8\u05e9\u05d9\u05dd</span><br>'
                for m in item_ministries:
                    m_name = m.get("name_he", m.get("name", ""))
                    item_html += f'<span style="display:inline-block;background:#eff6ff;border:1px solid #bfdbfe;border-radius:12px;padding:3px 10px;margin:2px;font-size:11px;color:#1e40af;">{m_name}</span>'
                item_html += '</div>'
                has_content = True

        # Per-item FTA
        if using_enriched:
            item_fta = c.get("fta")
            if item_fta and item_fta.get("eligible"):
                agreement = item_fta.get("agreement", item_fta.get("agreement_name", ""))
                pref_rate = item_fta.get("preferential", item_fta.get("preferential_rate", ""))
                origin_proof = item_fta.get("origin_proof", item_fta.get("documents_needed", ""))
                item_html += f'<div style="margin-top:6px;padding:6px 10px;background:#f0fdf4;border:1px solid #bbf7d0;border-radius:4px;font-size:12px;">'
                item_html += f'<strong style="color:#166534;">FTA:</strong> {agreement} '
                item_html += f'<span style="color:#166534;font-weight:bold;">\u05de\u05db\u05e1 \u05de\u05d5\u05e4\u05d7\u05ea {pref_rate}</span>'
                if origin_proof:
                    item_html += f' <span style="color:#555;">| \u05ea\u05e2\u05d5\u05d3\u05d4: {origin_proof}</span>'
                item_html += '</div>'
                has_content = True

        if not has_content:
            continue

        html += f"""
<tr><td style="padding:4px 30px;">
  <table width="100%" cellpadding="0" cellspacing="0"
         style="border:1px solid #e0e0e0;border-radius:4px;overflow:hidden;">
  <tr><td style="padding:10px 16px;background:#f8faff;border-bottom:1px solid #e5e7eb;">
    <span style="font-family:'Courier New',monospace;font-weight:bold;color:{_RPA_BLUE};font-size:13px;">{get_israeli_hs_format(hs)}</span>
    <span style="font-size:12px;color:#555;margin-right:8px;">{item_name}</span>
  </td></tr>
  <tr><td style="padding:12px 16px;">
    {item_html}
  </td></tr>
  </table>
</td></tr>"""

    return html


def _cls_fta_benefits(intelligence, fta):
    """FTA benefits section."""
    intel_fta_list = []
    if intelligence:
        for k, v in intelligence.items():
            if isinstance(v, dict):
                fta_data = v.get("fta")
                if isinstance(fta_data, dict) and fta_data.get("eligible"):
                    intel_fta_list.append(fta_data)
    ai_fta = [f for f in (fta or []) if f.get("eligible")]

    if not intel_fta_list and not ai_fta:
        return ""

    html = _cls_section_title("\u05d4\u05d8\u05d1\u05d5\u05ea \u05e1\u05d7\u05e8 (FTA)", color="#166534")

    fta_items = intel_fta_list if intel_fta_list else ai_fta
    for ft in fta_items:
        agreement = ft.get("agreement_name_he") or ft.get("agreement_name", ft.get("agreement", ""))
        pref_rate = ft.get("preferential_rate", ft.get("preferential", ""))
        country = ft.get("origin_country", ft.get("country", ""))
        origin_proof = ft.get("origin_proof", ft.get("documents_needed", ""))
        cumulation = ft.get("cumulation", "")
        legal = ft.get("legal_basis", "")

        html += f"""
<tr><td style="padding:4px 30px;">
  <table width="100%" cellpadding="0" cellspacing="0"
         style="border:1px solid #bbf7d0;border-radius:4px;overflow:hidden;background:#f0fdf4;">
  <tr>
    <td style="padding:12px 16px;vertical-align:middle;">
      <strong style="color:#166534;font-size:14px;">{country}</strong>
      <span style="color:#555;font-size:13px;"> &mdash; {agreement}</span>
    </td>"""
        if pref_rate:
            html += f"""
    <td style="padding:12px 16px;text-align:left;vertical-align:middle;">
      <span style="display:inline-block;background:{_COLOR_OK};color:#fff;font-weight:bold;padding:4px 14px;border-radius:12px;font-size:13px;">\u05de\u05db\u05e1 {pref_rate}</span>
    </td>"""
        html += """
  </tr>"""
        details = []
        if origin_proof:
            details.append(f'\u05d4\u05d5\u05db\u05d7\u05ea \u05de\u05e7\u05d5\u05e8: <strong>{origin_proof}</strong>')
        if cumulation:
            details.append(f'\u05e6\u05d1\u05d9\u05e8\u05d4: {cumulation}')
        if legal:
            details.append(f'\u05d1\u05e1\u05d9\u05e1 \u05de\u05e9\u05e4\u05d8\u05d9: {legal}')
        if details:
            html += '  <tr><td colspan="2" style="padding:6px 16px 12px;border-top:1px solid #bbf7d0;">'
            for d in details:
                html += f'<div style="font-size:12px;color:#555;margin-bottom:2px;">{d}</div>'
            html += '</td></tr>'
        html += """  </table>
</td></tr>"""

    return html


def _cls_risk(risk):
    """Risk section."""
    if not risk or risk.get("level") not in ("\u05d2\u05d1\u05d5\u05d4", "\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9"):
        return ""
    level = risk.get("level", "")
    risk_color = _COLOR_ERR if level == "\u05d2\u05d1\u05d5\u05d4" else _COLOR_WARN
    risk_border = "#fca5a5" if level == "\u05d2\u05d1\u05d5\u05d4" else "#fde68a"

    html = _cls_section_title(f"\u05e1\u05d9\u05db\u05d5\u05df: {level}", color=risk_color)

    for item in risk.get("items", []):
        html += f"""
<tr><td style="padding:4px 30px;">
  <table width="100%" cellpadding="0" cellspacing="0"
         style="border:1px solid {risk_border};border-radius:4px;overflow:hidden;">
  <tr><td style="padding:10px 16px;">
    <strong style="color:#333;">{item.get("item", "")}</strong>
    <div style="font-size:13px;color:#555;margin-top:4px;">{item.get("issue", "")}</div>
  </td></tr>
  </table>
</td></tr>"""

    return html


def _cls_smart_questions(smart_q, classifications):
    """Section 7: Smart questions when confidence is low."""
    if not smart_q or not SMART_QUESTIONS_AVAILABLE:
        return ""
    try:
        item_desc = ""
        if classifications:
            item_desc = classifications[0].get("item", "")
        questions_html = format_questions_html(smart_q, item_description=item_desc)
        if questions_html:
            return f"""
<!-- SMART QUESTIONS -->
<tr><td style="padding:10px 30px;">
  {questions_html}
</td></tr>"""
    except Exception:
        pass
    return ""


def _cls_original_email(original_email_body):
    """Original email quote."""
    if not original_email_body or len(original_email_body.strip()) <= 10:
        return ""
    import html as html_mod
    quoted = html_mod.escape(original_email_body.strip())
    if len(quoted) > 2000:
        quoted = quoted[:2000] + "..."
    return f"""
<!-- ORIGINAL EMAIL -->
<tr><td style="padding:10px 30px;">
  <table width="100%" cellpadding="0" cellspacing="0">
  <tr><td style="font-size:12px;color:#888;padding-bottom:6px;">\u05d4\u05d5\u05d3\u05e2\u05d4 \u05de\u05e7\u05d5\u05e8\u05d9\u05ea:</td></tr>
  <tr><td style="padding:10px 14px;border-right:3px solid #ccc;background:#f9f9f9;border-radius:4px;font-size:13px;color:#555;direction:rtl;">
    {quoted}
  </td></tr>
  </table>
</td></tr>"""


def _cls_footer():
    """Section 8: Branded footer with logo, timestamp, disclaimer."""
    from datetime import datetime, timezone
    if _to_israel_time:
        now = _to_israel_time(datetime.now(timezone.utc))
    else:
        now = datetime.now(timezone.utc)
    timestamp = f"{now.day:02d}/{now.month:02d}/{now.year} {now.hour:02d}:{now.minute:02d} IL"

    return f"""
<!-- FOOTER -->
<tr><td style="padding:20px 30px;border-top:2px solid #e0e0e0;background:#f8f9fa;">
  <table width="100%" cellpadding="0" cellspacing="0">
  <tr>
    <td style="padding-bottom:10px;">
      <span style="display:inline-block;vertical-align:middle;width:32px;height:32px;line-height:32px;text-align:center;background:#1e3a5f;color:#ffffff;font-size:12px;font-weight:bold;border-radius:4px;font-family:Georgia,'Times New Roman',serif;">RCB</span>
      <span style="display:inline-block;vertical-align:middle;padding-left:8px;">
        <span style="font-size:12px;font-weight:bold;color:{_RPA_BLUE};">RCB &#8212; AI Customs Broker</span><br>
        <span style="font-size:11px;color:#888;">R.P.A. PORT LTD</span>
        <span style="color:#ccc;margin:0 6px;">|</span>
        <span style="font-size:11px;color:{_RPA_BLUE};">rcb@rpa-port.co.il</span>
      </span>
    </td>
  </tr>
  <tr>
    <td style="font-size:10px;color:#999;padding-top:6px;border-top:1px solid #eee;">
      {timestamp}
    </td>
  </tr>
  <tr>
    <td style="font-size:10px;color:#aaa;padding-top:8px;direction:rtl;text-align:right;line-height:1.5;">
      &#9888; \u05d4\u05de\u05dc\u05e6\u05d4 \u05e8\u05d0\u05e9\u05d5\u05e0\u05d9\u05ea \u05d1\u05dc\u05d1\u05d3. \u05d9\u05e9 \u05dc\u05d0\u05de\u05ea \u05e2\u05dd \u05e2\u05de\u05d9\u05dc \u05de\u05db\u05e1 \u05de\u05d5\u05e1\u05de\u05da.<br>
      \u05d3\u05d5\u05f4\u05d7 \u05d6\u05d4 \u05d4\u05d5\u05e4\u05e7 \u05d1\u05d0\u05d5\u05e4\u05df \u05d0\u05d5\u05d8\u05d5\u05de\u05d8\u05d9 \u05d5\u05d0\u05d9\u05e0\u05d5 \u05de\u05d4\u05d5\u05d5\u05d4 \u05d9\u05d9\u05e2\u05d5\u05e5 \u05e8\u05e9\u05de\u05d9.
    </td>
  </tr>
  </table>
</td></tr>
"""


def build_classification_email(results, sender_name, invoice_validation=None, tracking_code=None, invoice_data=None, enriched_items=None, original_email_body=None):
    """Build HTML email report â€” table-based, Outlook-safe, RTL, branded.

    Session 40: Redesigned to match tracker_email.py design language.
    8 sections: header, result table, regulatory, cross-reference,
    customs value, justification, smart questions, footer.
    """
    classifications = results.get("agents", {}).get("classification", {}).get("classifications", [])
    regulatory = results.get("agents", {}).get("regulatory", {}).get("regulatory", [])
    fta = results.get("agents", {}).get("fta", {}).get("fta", [])
    risk = results.get("agents", {}).get("risk", {}).get("risk", {})
    synthesis = results.get("synthesis", "")
    ministry_routing = results.get("ministry_routing", {})
    intelligence = results.get("intelligence", {})

    if not tracking_code:
        tracking_code = generate_tracking_code()

    card_items = enriched_items if enriched_items else classifications
    _using_enriched = bool(enriched_items)

    # Use tracker_email wrappers if available, else local fallback
    open_fn = _html_open if _html_open else _cls_html_open
    close_fn = _html_close if _html_close else _cls_html_close

    # â”€â”€ Compose email from section builders â”€â”€
    html = open_fn()
    html += _cls_header(tracking_code)
    html += _cls_shipment_info(invoice_data)
    html += _cls_invoice_validation(invoice_validation)

    # Quality Gate banner
    audit_banner = results.get("audit", {}).get("warning_banner", "")
    if audit_banner:
        html += f'<tr><td style="padding:10px 30px;">{audit_banner}</td></tr>'

    html += _cls_synthesis(synthesis)
    html += _cls_result_table(card_items, _using_enriched)
    html += _cls_regulatory(ministry_routing, regulatory)
    html += _cls_cross_reference(results, card_items)
    html += _cls_customs_value(invoice_data, results)
    html += _cls_justification_details(card_items, _using_enriched)
    html += _cls_fta_benefits(intelligence, fta)
    html += _cls_risk(risk)
    html += _cls_smart_questions(results.get("smart_questions", []), classifications)
    html += _cls_original_email(original_email_body)
    html += _cls_footer()
    html += close_fn()

    return html

def build_excel_report(results, enriched_items=None):
    """Build Excel report with enriched item data when available."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "×¡×™×›×•×"
        ws['A1'] = "×“×•×´×— ×¡×™×•×•×’ RCB"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A3'] = results.get("synthesis", "")

        ws2 = wb.create_sheet("×¡×™×•×•×’×™×")
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        if enriched_items:
            # Enhanced columns
            headers = ["×©×•×¨×”", "×ª×™××•×¨ ×¤×¨×™×˜", "××•×›×¨", "×§×•× ×”", "×›××•×ª", "××—×™×¨ ×œ×™×—×™×“×”",
                        "××¨×¥ ××§×•×¨", "×§×•×“ HS", "×ª×™××•×¨ ×ª×¢×¨×™×£", "××›×¡", "××¡ ×§× ×™×™×”",
                        "××¢\"×", "×•×“××•×ª", "×¡×˜×˜×•×¡ ××™××•×ª", "××©×¨×“×™× × ×“×¨×©×™×", "× ×™××•×§"]
            for i, h in enumerate(headers, 1):
                cell = ws2.cell(1, i, h)
                cell.font = header_font
                cell.fill = header_fill
            for row, c in enumerate(enriched_items, 2):
                ws2.cell(row, 1, c.get("line_number", ""))
                ws2.cell(row, 2, c.get("description", ""))
                ws2.cell(row, 3, c.get("seller", ""))
                ws2.cell(row, 4, c.get("buyer", ""))
                ws2.cell(row, 5, c.get("quantity", ""))
                ws2.cell(row, 6, c.get("unit_price", ""))
                ws2.cell(row, 7, c.get("origin_country", ""))
                ws2.cell(row, 8, get_israeli_hs_format(c.get("hs_code", "")))
                ws2.cell(row, 9, c.get("tariff_text_he", ""))
                ws2.cell(row, 10, c.get("duty_rate", ""))
                ws2.cell(row, 11, c.get("purchase_tax_display", "×œ× ×—×œ"))
                ws2.cell(row, 12, c.get("vat_rate", ""))
                ws2.cell(row, 13, c.get("confidence", ""))
                v_status = c.get('verification_status', '')
                if v_status == 'official':
                    validation_status = "××•××ª ×¨×©××™×ª"
                elif v_status == 'verified':
                    validation_status = "××•××ª"
                elif c.get('hs_corrected'):
                    validation_status = f"×ª×•×§×Ÿ ×-{get_israeli_hs_format(c.get('original_hs_code', ''))}"
                else:
                    validation_status = v_status
                ws2.cell(row, 14, validation_status)
                # Ministries
                ministries = c.get("ministries", [])
                ministry_names = ", ".join(m.get("name_he", m.get("name", "")) for m in ministries) if ministries else ""
                ws2.cell(row, 15, ministry_names)
                ws2.cell(row, 16, c.get("reasoning", ""))
        else:
            # Fallback: original 8-column format
            headers = ["×¤×¨×™×˜", "×§×•×“ HS", "××›×¡", "××¡ ×§× ×™×™×”", "××¢\"×", "×•×“××•×ª", "×¡×˜×˜×•×¡ ××™××•×ª", "× ×™××•×§"]
            for i, h in enumerate(headers, 1):
                ws2.cell(1, i, h).font = Font(bold=True)
            for row, c in enumerate(results.get("agents", {}).get("classification", {}).get("classifications", []), 2):
                ws2.cell(row, 1, c.get("item", ""))
                ws2.cell(row, 2, get_israeli_hs_format(c.get("hs_code", "")))
                ws2.cell(row, 3, c.get("duty_rate", ""))
                pt = c.get("purchase_tax", {})
                pt_text = pt.get("rate_he", "×œ× ×—×œ") if isinstance(pt, dict) else "×œ× ×—×œ"
                ws2.cell(row, 4, pt_text)
                ws2.cell(row, 5, c.get("vat_rate", ""))
                ws2.cell(row, 6, c.get("confidence", ""))
                v_status = c.get('verification_status', '')
                if v_status == 'official':
                    validation_status = "××•××ª ×¨×©××™×ª âœ“"
                elif v_status == 'verified':
                    validation_status = "××•××ª âœ“"
                elif c.get('hs_corrected'):
                    validation_status = f"×ª×•×§×Ÿ ×-{get_israeli_hs_format(c.get('original_hs_code', ''))}"
                elif c.get('hs_warning'):
                    validation_status = "×œ× × ××¦× ×‘×××’×¨"
                elif c.get('hs_validated') and c.get('hs_exact_match'):
                    validation_status = "××•××ª âœ“"
                elif c.get('hs_validated'):
                    validation_status = "×”×ª×××” ×—×œ×§×™×ª"
                else:
                    validation_status = ""
                ws2.cell(row, 7, validation_status)
                ws2.cell(row, 8, c.get("reasoning", ""))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return base64.b64encode(output.read()).decode('utf-8')
    except Exception as e:
        print(f"Excel error: {e}")
        return None


def _parse_invoice_fields_from_data(invoice_data):
    """Session 10: Parse invoice data into validation format"""
    fields = {}
    
    if invoice_data.get('seller'):
        fields['seller'] = invoice_data['seller']
    if invoice_data.get('buyer'):
        fields['buyer'] = invoice_data['buyer']
    if invoice_data.get('invoice_date'):
        fields['date'] = invoice_data['invoice_date']
    if invoice_data.get('total_value'):
        fields['price'] = invoice_data['total_value']
    if invoice_data.get('incoterms'):
        fields['terms'] = invoice_data['incoterms']
    if invoice_data.get('currency'):
        fields['currency'] = invoice_data['currency']
    
    # Check items for origin and description
    items = invoice_data.get('items', [])
    if items:
        origins = [i.get('origin_country') for i in items if i.get('origin_country')]
        if origins:
            fields['origin'] = origins[0]
        
        descriptions = [i.get('description') for i in items if i.get('description')]
        if descriptions:
            fields['description'] = ', '.join(descriptions[:3])
        
        quantities = [i.get('quantity') for i in items if i.get('quantity')]
        if quantities:
            fields['quantity'] = ', '.join(str(q) for q in quantities[:3])
    
    return fields


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  EMAIL ENRICHMENT: Merge invoice items + classifications + tariff
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _lookup_tariff_text(db, hs_code):
    """Fallback: look up tariff description from Firestore for a specific HS code."""
    if not db or not hs_code:
        return "", ""
    hs_clean = str(hs_code).replace(".", "").replace(" ", "").replace("/", "")
    for coll_name in ["tariff", "tariff_chapters"]:
        try:
            candidates = [hs_clean]
            if len(hs_clean) >= 6:
                candidates.append(f"{hs_clean[:4]}.{hs_clean[4:6]}.{hs_clean[6:]}")
            if len(hs_clean) >= 4:
                candidates.append(hs_clean[:4])
            for doc_id in candidates:
                doc = db.collection(coll_name).document(doc_id).get()
                if doc.exists:
                    data = doc.to_dict()
                    return (
                        data.get("description_he", data.get("title_he", "")),
                        data.get("description_en", data.get("title_en", ""))
                    )
        except Exception:
            continue
    return "", ""


def _enrich_results_for_email(results, invoice_data, db):
    """Merge invoice items (Agent 1) with classifications (Agent 2) by index.
    Add tariff text, ministry routing, and FTA data per item."""
    invoice_items = results.get("agents", {}).get("invoice", {}).get("items", [])
    if not invoice_items and isinstance(invoice_data, dict):
        invoice_items = invoice_data.get("items", [])
    classifications = results.get("agents", {}).get("classification", {}).get("classifications", [])
    ministry_routing = results.get("ministry_routing", {})
    fta_data = results.get("agents", {}).get("fta", {}).get("fta", [])
    free_import = results.get("free_import_order", {})
    ve_results = results.get("verification_engine", {})

    seller = (invoice_data or {}).get("seller", "")
    buyer = (invoice_data or {}).get("buyer", "")

    enriched = []
    count = max(len(invoice_items), len(classifications))
    if count == 0:
        return enriched

    for idx in range(count):
        inv = invoice_items[idx] if idx < len(invoice_items) and isinstance(invoice_items[idx], dict) else {}
        cls = classifications[idx] if idx < len(classifications) and isinstance(classifications[idx], dict) else {}

        hs_code = cls.get("hs_code", "")

        # Tariff text: prefer verification loop data, fallback Firestore lookup
        tariff_he = cls.get("official_description_he", "")
        tariff_en = cls.get("official_description_en", "")
        if not tariff_he and hs_code and db:
            try:
                tariff_he, tariff_en = _lookup_tariff_text(db, hs_code)
            except Exception:
                pass

        # Ministry routing for this HS code
        ministries = []
        routing = ministry_routing.get(hs_code, {})
        if isinstance(routing, dict):
            ministries = routing.get("ministries", [])

        # FTA match by origin country
        origin = inv.get("origin_country", "")
        item_fta = None
        for f in fta_data:
            if not f.get("eligible"):
                continue
            f_country = (f.get("country", "") or "").lower()
            if origin and (f_country == origin.lower() or f_country in origin.lower()):
                item_fta = f
                break

        # Purchase tax display
        pt = cls.get("purchase_tax", {})
        if isinstance(pt, dict) and pt.get("applies"):
            pt_display = pt.get("rate_he", pt.get("rate", "×—×œ"))
        elif isinstance(pt, str):
            pt_display = pt
        else:
            pt_display = "×œ× ×—×œ"

        enriched.append({
            "line_number": idx + 1,
            "description": inv.get("description", cls.get("item", "")),
            "quantity": inv.get("quantity", ""),
            "unit_price": inv.get("unit_price", ""),
            "total": inv.get("total", ""),
            "origin_country": origin,
            "seller": seller,
            "buyer": buyer,
            "hs_code": hs_code,
            "tariff_text_he": tariff_he,
            "tariff_text_en": tariff_en,
            "duty_rate": cls.get("duty_rate", ""),
            "purchase_tax": pt,
            "purchase_tax_display": pt_display,
            "vat_rate": cls.get("vat_rate", ""),
            "confidence": cls.get("confidence", ""),
            "reasoning": cls.get("reasoning", ""),
            "verification_status": cls.get("verification_status", ""),
            "hs_corrected": cls.get("hs_corrected", False),
            "original_hs_code": cls.get("original_hs_code", ""),
            "hs_warning": cls.get("hs_warning", ""),
            "hs_validated": cls.get("hs_validated", False),
            "hs_exact_match": cls.get("hs_exact_match", False),
            "ministries": ministries,
            "fta": item_fta,
            "ve_flags": (ve_results.get(hs_code, {}).get("flags", []) if ve_results else []),
            "ve_phase4": (ve_results.get(hs_code, {}).get("phase4", {}) if ve_results else {}),
            "cross_ref_adjustment": cls.get("cross_ref_adjustment"),
            "cross_ref_note": cls.get("cross_ref_note", ""),
            "item": cls.get("item", inv.get("description", "")),
            "item_description": cls.get("item_description", cls.get("item", inv.get("description", ""))),
            "justification": cls.get("justification"),
            "challenge": cls.get("challenge"),
            "cross_check_tier": cls.get("cross_check_tier", 0),
            "cross_check_note": cls.get("cross_check_note", ""),
        })

    return enriched


# =============================================================================
# PRE-CLASSIFY BYPASS: Skip AI pipeline for high-confidence known products
# =============================================================================

def _try_pre_classify_bypass(db, doc_text, gemini_key=None, api_key=None):
    """
    Attempt to classify using ONLY pre_classify (Firestore lookups, zero AI).
    Returns a full result dict (same format as run_full_classification) or None.

    Only succeeds when:
    - PRE_CLASSIFY_BYPASS_ENABLED is True
    - pre_classify returns a candidate with confidence >= PRE_CLASSIFY_BYPASS_THRESHOLD
    - The top candidate has a valid HS code

    Session 48: Added api_key param so Claude fallback works when Gemini 429.
    """
    if not PRE_CLASSIFY_BYPASS_ENABLED or not INTELLIGENCE_AVAILABLE:
        return None

    # We need a minimal invoice extraction first (Gemini Flash â€” cheap, Claude fallback)
    print("  [BYPASS] Checking pre-classify bypass...")
    invoice = run_document_agent(api_key, doc_text, gemini_key=gemini_key)
    if not isinstance(invoice, dict):
        invoice = {"items": [{"description": doc_text[:500]}]}
    items = invoice.get("items") or [{"description": doc_text[:500]}]
    if not isinstance(items, list):
        items = [{"description": doc_text[:500]}]

    origin = items[0].get("origin_country", "") if items and isinstance(items[0], dict) else ""
    seller_name = invoice.get("seller", "")

    # Run pre_classify on first item
    best_candidate = None
    best_pc_result = None
    for item in items[:3]:
        if not isinstance(item, dict):
            continue
        desc = item.get("description", "")
        if not desc or not _is_product_description(desc):
            continue
        item_origin = item.get("origin_country", origin)
        pc_result = pre_classify(db, desc, item_origin, seller_name=seller_name)
        if not isinstance(pc_result, dict):
            continue
        candidates = pc_result.get("candidates", [])
        if candidates and candidates[0].get("confidence", 0) >= PRE_CLASSIFY_BYPASS_THRESHOLD:
            top = candidates[0]
            hs = str(top.get("hs_code", "")).replace(".", "").replace("/", "").replace(" ", "")
            if _is_valid_hs(hs):
                if not best_candidate or top["confidence"] > best_candidate["confidence"]:
                    best_candidate = top
                    best_pc_result = pc_result

    if not best_candidate:
        print("  [BYPASS] No high-confidence match â€” proceeding to AI pipeline")
        return None

    hs_code = best_candidate["hs_code"]
    confidence = best_candidate["confidence"]
    source = best_candidate.get("source", "unknown")
    print(f"  [BYPASS] HIGH-CONFIDENCE MATCH: HS {hs_code} ({confidence}% from {source})")
    print(f"  [BYPASS] Skipping 6-agent AI pipeline â€” $0.00 AI cost")

    # Build classifications in the standard format
    classifications = []
    for item in items[:10]:
        if not isinstance(item, dict):
            continue
        desc = item.get("description", "")
        classifications.append({
            "item": desc[:100],
            "item_description": desc,
            "hs_code": hs_code,
            "confidence": "×’×‘×•×”×”",
            "reasoning": f"Pre-classify bypass ({source}, {confidence}% confidence)",
            "reasoning_he": f"×¡×™×•×•×’ ××”×™×¨ ××××’×¨ ×”×™×“×¢ ({source}, {confidence}% ×‘×™×˜×—×•×Ÿ)",
            "hs_description": best_candidate.get("description", ""),
            "duty_rate": best_candidate.get("duty_rate", ""),
        })

    # Validate HS codes
    validated = validate_and_correct_classifications(db, classifications)

    # Free Import Order
    free_import_results = {}
    if INTELLIGENCE_AVAILABLE:
        try:
            fio = query_free_import_order(db, hs_code)
            if fio.get("found"):
                free_import_results[hs_code] = fio
        except Exception:
            pass

    # Ministry routing
    ministry_routing = {}
    if INTELLIGENCE_AVAILABLE:
        try:
            routing = route_to_ministries(db, hs_code, free_import_results.get(hs_code))
            if routing.get("ministries"):
                ministry_routing[hs_code] = routing
        except Exception:
            pass

    # Verification
    if VERIFICATION_AVAILABLE:
        try:
            validated = verify_all_classifications(db, validated, free_import_results=free_import_results)
            for c in validated:
                if c.get("verification_status") in ("official", "verified"):
                    learn_from_verification(db, c)
        except Exception:
            pass

    # Build synthesis from pre-classify data
    synthesis = (
        f"×¡×™×•×•×’ ××”×™×¨: ×”××¢×¨×›×ª ×–×™×”×ª×” ××ª ×”××•×¦×¨ ××××’×¨ ×”×™×“×¢ ×”×¤× ×™××™ "
        f"(××§×•×¨: {source}, ×¨××ª ×‘×™×˜×—×•×Ÿ: {confidence}%). "
        f"×§×•×“ HS: {hs_code}."
    )
    if best_candidate.get("duty_rate"):
        synthesis += f" ×©×™×¢×•×¨ ××›×¡: {best_candidate['duty_rate']}."
    if free_import_results:
        synthesis += " × ×‘×“×§ ××•×œ ×¦×• ×™×‘×•× ×—×•×¤×©×™."

    all_results = {
        "invoice": invoice,
        "classification": {"classifications": validated},
        "regulatory": {"regulatory": best_pc_result.get("regulatory", [])},
        "fta": {"fta": [best_pc_result.get("fta")] if best_pc_result.get("fta") else []},
        "risk": {"risk": {"level": "× ××•×š", "items": []}},
    }

    return {
        "success": True,
        "agents": all_results,
        "synthesis": synthesis,
        "invoice_data": invoice,
        "intelligence": {"bypass": best_pc_result},
        "document_validation": None,
        "free_import_order": free_import_results,
        "ministry_routing": ministry_routing,
        "parsed_documents": [],
        "smart_questions": [],
        "ambiguity": {},
        "tracker": None,
        "audit": {"action": "send", "classifications": validated, "warning_banner": None, "avg_confidence": confidence},
        "_engine": "pre_classify_bypass",
        "_bypass_confidence": confidence,
        "_bypass_source": source,
        "_cost": 0.001,  # Only Gemini Flash for invoice extraction
    }


def _link_classification_to_tracker(db, classification_doc_id, invoice_data, results, tracking_code):
    """Bridge: link a completed classification to an existing tracker deal.
    Looks up tracker_deals by BL or AWB from invoice data, writes back:
      - rcb_classification_id (the Firestore doc ID in rcb_classifications)
      - rcb_tracking_code (the human-readable RCB-XXXX code)
      - classification_hs_codes (list of {hs_code, description, confidence})
    """
    bl = invoice_data.get('bl_number', '').strip()
    awb = invoice_data.get('awb_number', '').strip()

    if not bl and not awb:
        return  # No shipment reference to match

    # Find the tracker deal
    deal_doc = None
    if bl:
        results_q = list(db.collection("tracker_deals")
                         .where("bol_number", "==", bl)
                         .where("status", "in", ["active", "pending", "completed"])
                         .limit(1).stream())
        if results_q:
            deal_doc = results_q[0]
    if not deal_doc and awb:
        results_q = list(db.collection("tracker_deals")
                         .where("awb_number", "==", awb)
                         .where("status", "in", ["active", "pending", "completed"])
                         .limit(1).stream())
        if results_q:
            deal_doc = results_q[0]

    if not deal_doc:
        print(f"  Bridge: no tracker deal found for BL={bl} AWB={awb}")
        return

    # Extract HS codes from classification results
    hs_codes = []
    try:
        classifications = (results.get("agents", {}).get("classification", {})
                          .get("classifications", []))
        for c in classifications:
            hs = c.get("hs_code", "")
            if hs:
                hs_codes.append({
                    "hs_code": hs,
                    "description": c.get("item", "")[:100],
                    "confidence": c.get("confidence", ""),
                })
    except Exception:
        pass

    # Write back to tracker deal
    from datetime import datetime, timezone
    deal_doc.reference.update({
        "rcb_classification_id": classification_doc_id,
        "rcb_tracking_code": tracking_code,
        "classification_hs_codes": hs_codes,
        "classification_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    })

    print(f"  Bridge: linked classification {tracking_code} â†’ deal {deal_doc.id} "
          f"({len(hs_codes)} HS codes)")


def process_and_send_report(access_token, rcb_email, to_email, subject, sender_name, raw_attachments, msg_id, get_secret_func, db, firestore, helper_graph_send, extract_text_func, email_body=None, internet_message_id=None):
    """Main: Extract, classify, validate, send ONE consolidated email.

    Sends a single email containing: ack + classification report + clarification (if needed).
    Uses internet_message_id for proper Outlook/Gmail threading.

    Session 47: to_email may be None for external senders with no @rpa-port.co.il in chain.
    In that case, classification runs normally but NO reply email is sent.
    """
    # Session 47: Reply suppression for external senders without team address
    _suppress_reply = not to_email
    if _suppress_reply:
        print("  ğŸ“¬ No reply target (external sender, no team address) â€” will classify but NOT send email")
    try:
        print(f"  ğŸ¤– Starting: {subject[:50]}")
        _cost_tracker.reset()  # Session 27: Reset per-run cost accumulator
        try:
            from lib.librarian import clear_search_cache
            clear_search_cache()  # Session 27: Clear search cache between runs
        except ImportError:
            pass

        api_key = get_secret_func('ANTHROPIC_API_KEY')
        if not api_key:
            print("  âŒ No API key")
            return False
        api_key = api_key.strip()  # Session 48: Safety â€” strip whitespace from Secret Manager

        # Session 48+CRIT-2: Reset Gemini quota flag at start of each run
        global _gemini_quota_exhausted_at
        _gemini_quota_exhausted_at = 0

        # Session 15: Get Gemini key for cost-optimized agents
        gemini_key = None
        try:
            gemini_key = get_secret_func('GEMINI_API_KEY')
            if gemini_key:
                gemini_key = gemini_key.strip()  # Session 48: Safety strip
                print("  ğŸ”‘ Gemini key loaded (multi-model mode)")
            else:
                print("  â„¹ï¸ No Gemini key - all agents will use Claude")
        except Exception:
            print("  â„¹ï¸ Gemini key not configured - all agents will use Claude")
        
        print("  ğŸ“„ Extracting text...")
        doc_text = extract_text_func(raw_attachments, email_body=email_body,
                                     gemini_key=gemini_key, anthropic_key=api_key)
        if not doc_text or len(doc_text) < 50:
            print("  âš ï¸ No text â€” sending extraction failure notification")
            _fail_html = (
                '<div style="font-family:\'Segoe UI\',Arial,Helvetica,sans-serif;max-width:680px;margin:0 auto;direction:rtl;background:#f0f2f5;padding:0">'
                '<div style="background:linear-gradient(135deg,#0f2439 0%,#1a3a5c 50%,#245a8a 100%);color:#fff;padding:24px 30px 20px;border-radius:12px 12px 0 0">'
                '<h1 style="margin:0;font-size:20px;font-weight:700">×“×•×´×— ×¡×™×•×•×’ ××›×¡</h1>'
                '<p style="margin:4px 0 0 0;font-size:13px;opacity:0.8">RCB â€” AI Customs Broker</p>'
                '</div>'
                '<div style="background:#ffffff;padding:28px 30px;border-left:1px solid #e0e0e0;border-right:1px solid #e0e0e0">'
                '<div style="background:#fef2f2;border:2px solid #fca5a5;border-radius:10px;padding:20px">'
                '<h3 style="color:#991b1b;margin:0;font-size:16px">âš ï¸ ×œ× ×”×¦×œ×—× ×• ×œ×§×¨×•× ××ª ×”×§×‘×¦×™× ×”××¦×•×¨×¤×™×</h3>'
                '<p style="color:#555;margin:10px 0;font-size:14px">×§×™×‘×œ× ×• ××ª ×”××™×™×œ ×©×œ×š ××š ×œ× ×”×¦×œ×—× ×• ×œ×—×œ×¥ ×˜×§×¡×˜ ××”×§×‘×¦×™×. ×¡×™×‘×•×ª ××¤×©×¨×™×•×ª:</p>'
                '<ul style="color:#555;margin:6px 0;padding-right:20px;font-size:13px;line-height:1.8">'
                '<li>×§×•×‘×¥ ×¡×¨×•×§ (×ª××•× ×” ×‘×œ×‘×“, ×œ×œ× ×˜×§×¡×˜)</li>'
                '<li>PDF ××•×’×Ÿ ×‘×¡×™×¡××”</li>'
                '<li>×¤×•×¨××˜ ×œ× × ×ª××š (×œ××©×œ: .zip, .rar)</li>'
                '</ul>'
                '<p style="color:#333;font-weight:600;margin:14px 0 6px;font-size:14px">××” ×œ×¢×©×•×ª:</p>'
                '<ul style="color:#555;margin:4px 0;padding-right:20px;font-size:13px;line-height:1.8">'
                '<li>×©×œ×—×• PDF ××—×•×œ×œ ×××—×©×‘ (×œ× ×¡×¨×™×§×”)</li>'
                '<li>××• ×›×ª×‘×• ××ª ×ª×™××•×¨ ×”××•×¦×¨×™× ×‘×’×•×£ ×”××™×™×œ</li>'
                '</ul>'
                '</div></div>'
                '<div style="background:#f8faff;padding:16px 30px;border-top:1px solid #e0e0e0;border-radius:0 0 12px 12px;border-left:1px solid #e0e0e0;border-right:1px solid #e0e0e0">'
                '<p style="font-size:10px;color:#aaa;margin:0;line-height:1.5">âš ï¸ ×”××œ×¦×” ×¨××©×•× ×™×ª ×‘×œ×‘×“. ×™×© ×œ×××ª ×¢× ×¢××™×œ ××›×¡ ××•×¡××š. | RCB â€” rcb@rpa-port.co.il</p>'
                '</div></div>'
            )
            if not _suppress_reply:
                try:
                    helper_graph_send(access_token, rcb_email, to_email, subject,
                                      _fail_html, msg_id, internet_message_id=internet_message_id)
                except Exception:
                    pass
            try:
                import hashlib as _hl_f
                _safe_f = _hl_f.md5(msg_id.encode()).hexdigest()
                db.collection("rcb_processed").document(_safe_f).update({"type": "extraction_failed"})
            except Exception:
                pass
            return False
        
        print(f"  ğŸ“ {len(doc_text)} chars")

        # â”€â”€ PRE-CLASSIFY BYPASS: Skip AI if known product with high confidence â”€â”€
        results = None
        if PRE_CLASSIFY_BYPASS_ENABLED:
            try:
                results = _try_pre_classify_bypass(db, doc_text, gemini_key=gemini_key, api_key=api_key)
                if results:
                    print(f"  âœ… PRE-CLASSIFY BYPASS: saved ~$0.05 (engine={results.get('_engine')})")
            except Exception as bp_err:
                print(f"  âš ï¸ Pre-classify bypass error: {bp_err}")
                results = None

        # Session 48: Fetch OpenAI key early â€” needed for Agent 1 triple fallback
        openai_key = None
        try:
            openai_key = get_secret_func('OPENAI_API_KEY')
            if openai_key:
                openai_key = openai_key.strip()
        except Exception:
            pass

        # â”€â”€ TOOL-CALLING ENGINE: Single AI call with tools â”€â”€
        if not results and TOOL_CALLING_AVAILABLE:
            try:
                results = tool_calling_classify(api_key, doc_text, db, gemini_key=gemini_key)
                if not results or not results.get("success"):
                    print("  âš ï¸ Tool-calling returned no result, falling back to pipeline")
                    results = None
            except Exception as tc_err:
                print(f"  âš ï¸ Tool-calling error: {tc_err}, falling back to pipeline")
                results = None

        # â”€â”€ FULL PIPELINE FALLBACK: 6-agent sequential classification â”€â”€
        if not results:
            results = run_full_classification(api_key, doc_text, db, gemini_key=gemini_key, openai_key=openai_key)
        if not results.get('success'):
            print(f"  âŒ Classification failed â€” sending pipeline error notification")
            _err_html = (
                '<div style="font-family:\'Segoe UI\',Arial,Helvetica,sans-serif;max-width:680px;margin:0 auto;direction:rtl;background:#f0f2f5;padding:0">'
                '<div style="background:linear-gradient(135deg,#0f2439 0%,#1a3a5c 50%,#245a8a 100%);color:#fff;padding:24px 30px 20px;border-radius:12px 12px 0 0">'
                '<h1 style="margin:0;font-size:20px;font-weight:700">×“×•×´×— ×¡×™×•×•×’ ××›×¡</h1>'
                '<p style="margin:4px 0 0 0;font-size:13px;opacity:0.8">RCB â€” AI Customs Broker</p>'
                '</div>'
                '<div style="background:#ffffff;padding:28px 30px;border-left:1px solid #e0e0e0;border-right:1px solid #e0e0e0">'
                '<div style="background:#fef2f2;border:2px solid #fca5a5;border-radius:10px;padding:20px">'
                '<h3 style="color:#991b1b;margin:0;font-size:16px">âš ï¸ ×”×¡×™×•×•×’ ×œ× ×”×•×©×œ×</h3>'
                '<p style="color:#555;margin:10px 0;font-size:14px">×§×™×‘×œ× ×• ××ª ×”××™×™×œ ×•×§×¨×× ×• ××ª ×”×§×‘×¦×™× ×‘×”×¦×œ×—×”, ××š × ×ª×§×œ× ×• ×‘×©×’×™××” ×˜×›× ×™×ª ×‘××”×œ×š ×”×¡×™×•×•×’.</p>'
                '<p style="color:#333;font-weight:600;margin:14px 0 6px;font-size:14px">××” ×œ×¢×©×•×ª:</p>'
                '<ul style="color:#555;margin:4px 0;padding-right:20px;font-size:13px;line-height:1.8">'
                '<li><strong>×©×œ×—×• ××ª ×”××™×™×œ ×©×•×‘</strong> â€” ×”××¢×¨×›×ª ×ª× ×¡×” ×©× ×™×ª</li>'
                '<li>×× ×”×‘×¢×™×” ×—×•×–×¨×ª, ×¤× ×• ×œ×¦×•×•×ª ×‘- rcb@rpa-port.co.il</li>'
                '</ul>'
                '</div></div>'
                '<div style="background:#f8faff;padding:16px 30px;border-top:1px solid #e0e0e0;border-radius:0 0 12px 12px;border-left:1px solid #e0e0e0;border-right:1px solid #e0e0e0">'
                '<p style="font-size:10px;color:#aaa;margin:0;line-height:1.5">âš ï¸ ×”××œ×¦×” ×¨××©×•× ×™×ª ×‘×œ×‘×“. ×™×© ×œ×××ª ×¢× ×¢××™×œ ××›×¡ ××•×¡××š. | RCB â€” rcb@rpa-port.co.il</p>'
                '</div></div>'
            )
            if not _suppress_reply:
                try:
                    helper_graph_send(access_token, rcb_email, to_email, subject,
                                      _err_html, msg_id, internet_message_id=internet_message_id)
                except Exception:
                    pass
            try:
                import hashlib as _hl_p
                _safe_p = _hl_p.md5(msg_id.encode()).hexdigest()
                db.collection("rcb_processed").document(_safe_p).update({"type": "pipeline_error"})
            except Exception:
                pass
            return False

        # â”€â”€ SESSION 27: THREE-WAY CROSS-CHECK â”€â”€
        if CROSS_CHECK_ENABLED and CROSS_CHECK_AVAILABLE:
            try:
                classifications = (
                    results.get("agents", {}).get("classification", {}).get("classifications", [])
                )
                invoice_data_cc = results.get("invoice_data", {})
                items_cc = invoice_data_cc.get("items", []) if isinstance(invoice_data_cc, dict) else []
                item_descs = [
                    it.get("description", "") for it in items_cc if isinstance(it, dict)
                ]
                origin_cc = ""
                if item_descs and items_cc:
                    origin_cc = items_cc[0].get("origin_country", "") if isinstance(items_cc[0], dict) else ""

                openai_key = None
                try:
                    openai_key = get_secret_func("OPENAI_API_KEY")
                except Exception:
                    pass

                if classifications:
                    cc_results = cross_check_all_items(
                        classifications,
                        item_descs,
                        origin_cc,
                        api_key,
                        gemini_key,
                        openai_key,
                        db=db,
                    )
                    if cc_results:
                        results["cross_check"] = cc_results
                        cc_summary = get_cross_check_summary(cc_results)
                        if cc_summary:
                            existing_synthesis = results.get("synthesis", "")
                            results["synthesis"] = existing_synthesis + "\n\n" + cc_summary
                            print(f"  ğŸ” Cross-check: {len(cc_results)} items verified")
                        # Apply confidence adjustments
                        for i, cc in enumerate(cc_results):
                            adj = cc.get("confidence_adjustment", 0)
                            if i < len(classifications):
                                if adj != 0:
                                    _apply_confidence_adjustment(classifications[i], adj)
                                classifications[i]["cross_check_tier"] = cc.get("tier", 0)
                                classifications[i]["cross_check_note"] = cc.get("learning_note", "")
            except Exception as cc_err:
                print(f"  âš ï¸ Cross-check error: {cc_err}")

        # â”€â”€ HIGH-5: Apply cross-reference confidence adjustments â”€â”€
        try:
            _cr_cls = results.get("agents", {}).get("classification", {}).get("classifications", [])
            for c in _cr_cls:
                adj = c.get("cross_ref_adjustment", 0)
                if adj:
                    _apply_confidence_adjustment(c, adj)
        except Exception:
            pass

        # â”€â”€ SESSION 27: JUSTIFICATION + CHALLENGE (Assignment 11) â”€â”€
        try:
            from lib.justification_engine import (
                build_justification_chain,
                challenge_classification,
                save_knowledge_gaps,
                generate_pupil_questions,
            )

            classifications_j = (
                results.get("agents", {}).get("classification", {}).get("classifications", [])
            )

            all_gaps = []
            for idx, cls_item in enumerate(classifications_j):
                item_hs = cls_item.get("hs_code", "")
                item_desc = cls_item.get("item_description", cls_item.get("item", ""))[:500]
                if not item_hs:
                    continue

                # 1. Build justification chain
                justification = build_justification_chain(db, item_hs, item_desc)
                cls_item["justification"] = justification
                cls_item["legal_strength"] = justification["legal_strength"]

                # 2. Challenge (devil's advocate)
                challenge = challenge_classification(
                    db, item_hs, item_desc,
                    primary_chapter=item_hs[:2],
                    gemini_key=gemini_key,
                    call_gemini_func=call_gemini,
                )
                cls_item["challenge"] = challenge

                # 3. Collect gaps
                all_gaps.extend(justification.get("gaps", []))

                # 4. Generate questions if confused
                conf = cls_item.get("confidence", "")
                is_low = conf in ("low", "none") or (isinstance(conf, (int, float)) and conf < 0.7)
                if is_low or not challenge["challenge_passed"]:
                    questions = generate_pupil_questions(
                        db, item_hs, item_desc, justification, challenge,
                    )
                    if questions:
                        cls_item["pupil_questions"] = len(questions)
                        print(f"    ? {len(questions)} questions for daily digest")

                print(
                    f"    Legal: {justification['legal_strength']} "
                    f"({justification['coverage_pct']}% sourced), "
                    f"challenge: {'PASSED' if challenge['challenge_passed'] else 'OPEN'}"
                )

            # 5. Save all gaps for nightly enrichment
            if all_gaps:
                save_knowledge_gaps(db, all_gaps, email_id=msg_id)
                print(f"  Gaps: {len(all_gaps)} knowledge gaps saved")

        except ImportError:
            pass
        except Exception as just_err:
            print(f"  Justification error (non-fatal): {just_err}")

        # Session 10: Validate invoice using Module 5
        invoice_validation = None
        if MODULES_AVAILABLE:
            print("  ğŸ“‹ Validating invoice (Module 5)...")
            invoice_data = results.get('invoice_data', {})
            validation_fields = _parse_invoice_fields_from_data(invoice_data)
            
            try:
                validation_result = validate_invoice(validation_fields)
                invoice_validation = {
                    'score': validation_result.score,
                    'is_valid': validation_result.is_valid,
                    'missing_fields': [FIELD_DEFINITIONS[f]['name_he'] for f in validation_result.missing_fields],
                    'fields_present': validation_result.fields_present,
                    'fields_required': validation_result.fields_required,
                }
                print(f"  ğŸ“Š Invoice score: {validation_result.score}/100 ({'âœ…' if validation_result.is_valid else 'âš ï¸'})")
            except Exception as ve:
                print(f"  âš ï¸ Validation error: {ve}")
        
        # Session 11: Build standardized English subject line FIRST (need tracking code for HTML)
        invoice_data = results.get('invoice_data', {})
        score = invoice_validation['score'] if invoice_validation else None
        
        # Determine status
        has_smart_questions = bool(results.get("smart_questions"))
        _items_count = len(results.get("agents", {}).get("classification", {}).get("classifications", []))
        if invoice_validation and invoice_validation['score'] < 70:
            status = "CLARIFICATION"
        elif has_smart_questions:
            status = "CLARIFICATION"
        elif _items_count == 0:
            # Session 47: No items extracted â€” ask sender for product details
            status = "CLARIFICATION"
        elif invoice_validation and invoice_validation['score'] >= 70:
            status = "FINAL"
        else:
            status = "ACK"
        
        # AI-generated subject line with full shipment context
        tracking_code = generate_tracking_code()
        try:
            from lib.rcb_helpers import generate_smart_subject
            subject_line = generate_smart_subject(
                gemini_key=gemini_key, db=db, invoice_data=invoice_data,
                status=status, tracking_code=tracking_code,
                email_type="classification"
            )
        except Exception as e:
            print(f"    âš ï¸ Smart subject failed ({e}), using original")
            subject_line, tracking_code = build_rcb_subject(invoice_data, status, score)
        print(f"  ğŸ·ï¸ Tracking: {tracking_code}")
        
        print("  ğŸ“‹ Building reports...")
        enriched_items = _enrich_results_for_email(results, invoice_data, db)
        html = build_classification_email(results, sender_name, invoice_validation, tracking_code, invoice_data,
                                           enriched_items=enriched_items, original_email_body=email_body)
        excel = build_excel_report(results, enriched_items=enriched_items)

        # â”€â”€ Build clarification section (merged into same email) â”€â”€
        clarification_sent = False
        clarification_html = ""

        # Session 47: If zero items were classified, ask sender for product details
        if _items_count == 0:
            clarification_html = (
                '<div dir="rtl" style="background:#fff3cd;border:2px solid #ffc107;'
                'border-radius:10px;padding:20px;margin-top:25px;font-family:Arial,sans-serif">'
                '<h2 style="color:#856404;margin:0 0 10px 0">\U0001f4cb \u05d1\u05e7\u05e9\u05d4 \u05dc\u05d4\u05d1\u05d4\u05e8\u05d4</h2>'
                '<p style="color:#856404;margin:8px 0;font-size:14px">'
                '\u05dc\u05d0 \u05d4\u05e6\u05dc\u05d7\u05ea\u05d9 \u05dc\u05d6\u05d4\u05d5\u05ea \u05e4\u05e8\u05d9\u05d8\u05d9\u05dd \u05d1\u05d7\u05e9\u05d1\u05d5\u05e0\u05d9\u05ea. \u05d0\u05e0\u05d0 \u05e9\u05dc\u05d7:</p>'
                '<ul style="color:#856404;margin:8px 0;padding-right:20px;font-size:14px;line-height:2">'
                '<li>\u05ea\u05d9\u05d0\u05d5\u05e8 \u05d4\u05e1\u05d7\u05d5\u05e8\u05d4</li>'
                '<li>\u05db\u05de\u05d5\u05ea \u05d5\u05d9\u05d7\u05d9\u05d3\u05d5\u05ea</li>'
                '<li>\u05e2\u05e8\u05da \u05e4\u05e8\u05d9\u05d8</li>'
                '</ul></div>'
            )
            clarification_sent = True
            print("  \U0001f4cb Clarification: items=0, asking sender for product details")

        if MODULES_AVAILABLE and invoice_validation and invoice_validation['score'] < 70:
            try:
                from lib.clarification_generator import (
                    generate_missing_docs_request,
                    DocumentType,
                    UrgencyLevel,
                )

                missing_docs = []
                field_to_doc = {
                    '××¨×¥ ×”××§×•×¨': DocumentType.CERTIFICATE_OF_ORIGIN,
                    '×¤×¨×˜×™ ××¨×™×–×•×ª': DocumentType.PACKING_LIST,
                    '××©×§×œ×™×': DocumentType.PACKING_LIST,
                    '×ª× ××™ ××›×¨': DocumentType.INVOICE,
                    '××—×™×¨': DocumentType.INVOICE,
                    '×ª×™××•×¨ ×”×˜×•×‘×™×Ÿ': DocumentType.INVOICE,
                }

                for field_name in invoice_validation.get('missing_fields', []):
                    if field_name in field_to_doc:
                        doc = field_to_doc[field_name]
                        if doc not in missing_docs:
                            missing_docs.append(doc)

                if missing_docs:
                    urgency = UrgencyLevel.HIGH if invoice_validation['score'] < 50 else UrgencyLevel.MEDIUM
                    clarification = generate_missing_docs_request(
                        missing_docs=missing_docs,
                        invoice_number=invoice_data.get('invoice_number') or invoice_data.get('invoice_date'),
                        supplier_name=invoice_data.get('seller'),
                        recipient_name=sender_name,
                        urgency=urgency,
                        sender_name="××¢×¨×›×ª RCB",
                    )
                    clarification_html = f'''<div style="background:#fff3cd;border:2px solid #ffc107;border-radius:10px;padding:20px;margin-top:25px">
                        <h2 style="color:#856404;margin:0 0 10px 0">ğŸ“‹ ×‘×§×©×” ×œ×”×©×œ××ª ××¡××›×™×</h2>
                        <pre style="white-space:pre-wrap;font-family:Arial,sans-serif;direction:rtl;text-align:right;color:#856404;margin:0">{clarification.body}</pre>
                    </div>'''
                    clarification_sent = True
                    print(f"  ğŸ“‹ Clarification section added (score: {invoice_validation['score']}/100)")

            except Exception as ce:
                print(f"  âš ï¸ Clarification generation error: {ce}")

        # â”€â”€ Assemble ONE consolidated email: ack + classification + clarification â”€â”€
        # Insert ack banner before the classification report
        name = sender_name.split('<')[0].strip()
        first_name = name.split()[0] if name and '@' not in name else ""
        # Try Hebrew name
        try:
            from lib.rcb_helpers import to_hebrew_name
            display_name = to_hebrew_name(first_name) if first_name else "×©×œ×•×"
        except Exception:
            display_name = first_name or "×©×œ×•×"

        hour = datetime.now().hour
        greeting = "×‘×•×§×¨ ×˜×•×‘" if 5 <= hour < 12 else "×¦×”×¨×™×™× ×˜×•×‘×™×" if 12 <= hour < 17 else "×¢×¨×‘ ×˜×•×‘" if 17 <= hour < 21 else "×œ×™×œ×” ×˜×•×‘"

        ack_banner = f'''<div dir="rtl" style="font-family:Arial,sans-serif;max-width:800px;margin:0 auto 15px auto;background:#d4edda;border:1px solid #28a745;border-radius:10px;padding:15px">
            <p style="margin:0"><strong>{greeting} {display_name},</strong></p>
            <p style="margin:5px 0 0 0">âœ… ×§×™×‘×œ×ª×™ ××ª ×”××¡××›×™× ×•×¢×™×‘×“×ª×™ ××•×ª×. ×œ×”×œ×Ÿ ×”×“×•"×— ×”××œ×:</p>
        </div>'''

        # Final HTML: insert ack banner + clarification INSIDE the HTML structure
        # The ack banner goes right after the header; clarification before footer.
        # Both must stay inside </body></html>.
        final_html = html

        # Insert ack banner after the first <tr> header section
        header_end = final_html.find('<!-- HEADER -->')
        if header_end > 0:
            # Find end of header row
            tr_end = final_html.find('</tr>', header_end)
            if tr_end > 0:
                tr_end = final_html.find('>', tr_end) + 1
                insert_pos = tr_end
            else:
                insert_pos = header_end
        else:
            # Fallback: insert after first <table> row
            insert_pos = final_html.find('</tr>')
            if insert_pos > 0:
                insert_pos = final_html.find('>', insert_pos) + 1
            else:
                insert_pos = 0
        ack_row = f'<tr><td style="padding:15px 30px 0 30px;">{ack_banner}</td></tr>'
        final_html = final_html[:insert_pos] + ack_row + final_html[insert_pos:]

        if clarification_html:
            clar_row = f'<tr><td style="padding:0 30px 15px 30px;">{clarification_html}</td></tr>'
            # Insert before footer (case-insensitive search)
            footer_pos = final_html.rfind('<!-- FOOTER -->')
            if footer_pos < 0:
                footer_pos = final_html.rfind('<!-- Footer -->')
            if footer_pos > 0:
                final_html = final_html[:footer_pos] + clar_row + final_html[footer_pos:]
            else:
                # Last resort: insert before closing </table></td></tr></table></body></html>
                close_pos = final_html.rfind('</body>')
                if close_pos > 0:
                    final_html = final_html[:close_pos] + clar_row + final_html[close_pos:]
                else:
                    final_html = final_html + clar_row

        attachments = []
        if excel:
            attachments.append({
                '@odata.type': '#microsoft.graph.fileAttachment',
                'name': f'RCB_{tracking_code}_{datetime.now().strftime("%Y%m%d")}.xlsx',
                'contentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'contentBytes': excel
            })

        for att in raw_attachments:
            if att.get('contentBytes'):
                attachments.append({
                    '@odata.type': '#microsoft.graph.fileAttachment',
                    'name': att.get('name', 'file'),
                    'contentType': att.get('contentType', 'application/octet-stream'),
                    'contentBytes': att.get('contentBytes')
                })

        print("  ğŸ“¤ Sending consolidated email...")
        _email_sent = False
        if _suppress_reply:
            print("  ğŸ“¬ Reply suppressed (external sender, no team address)")
            _email_sent = True  # Treat as "sent" for Firestore save flow
        elif helper_graph_send(access_token, rcb_email, to_email, subject_line, final_html, msg_id, attachments, internet_message_id=internet_message_id):
            print(f"  âœ… Sent to {to_email}")
            _email_sent = True
        if _email_sent:

            # Save to Firestore
            save_data = {
                "tracking_code": tracking_code,
                "subject": subject_line,
                "original_subject": subject,
                "to": to_email,
                "items": len(results.get("agents", {}).get("classification", {}).get("classifications", [])),
                "timestamp": firestore.SERVER_TIMESTAMP,
                "direction": invoice_data.get('direction', 'unknown'),
                "freight_type": invoice_data.get('freight_type', 'unknown'),
                "bl_number": invoice_data.get('bl_number', ''),
                "awb_number": invoice_data.get('awb_number', ''),
                "seller": invoice_data.get('seller', ''),
                "buyer": invoice_data.get('buyer', ''),
                "clarification_sent": clarification_sent,
            }

            if invoice_validation:
                save_data["invoice_score"] = invoice_validation['score']
                save_data["invoice_valid"] = invoice_validation['is_valid']
                save_data["missing_fields"] = invoice_validation['missing_fields']

            if has_smart_questions:
                ambiguity = results.get("ambiguity", {})
                save_data["smart_questions_count"] = len(results["smart_questions"])
                save_data["ambiguity_reason"] = ambiguity.get("reason", "")
                save_data["classification_ambiguous"] = True

            # Session 27: Cost tracking
            if COST_TRACKING_ENABLED:
                save_data["total_cost"] = round(_cost_tracker.total(), 4)

            # Session 27: Cross-check results
            if results.get("cross_check"):
                cc_tiers = [cc.get("tier", 0) for cc in results["cross_check"]]
                save_data["cross_check_tiers"] = cc_tiers
                save_data["cross_check_avg_tier"] = round(sum(cc_tiers) / len(cc_tiers), 1) if cc_tiers else 0

            cls_ref = db.collection("rcb_classifications").add(save_data)
            cls_doc_id = cls_ref[1].id if isinstance(cls_ref, tuple) else getattr(cls_ref, 'id', '')

            # Session 33: Bridge classification â†’ tracker deal
            # Look up tracker deal by BL/AWB and write classification ID + HS codes back
            try:
                _link_classification_to_tracker(
                    db, cls_doc_id, invoice_data, results, tracking_code)
            except Exception as bridge_err:
                print(f"  Bridge error (non-fatal): {bridge_err}")

            # Session 27 Assignment 12: Save structured classification report
            try:
                from lib.report_builder import build_report_document
                report_doc = build_report_document(
                    results, tracking_code, to_email, invoice_data, invoice_validation,
                )
                db.collection("classification_reports").add(report_doc)
                print(f"  Classification report saved to Firestore")
            except Exception as rpt_err:
                print(f"  Report save error (non-fatal): {rpt_err}")

            # Learn from classification
            if ENRICHMENT_AVAILABLE:
                try:
                    enrichment = create_enrichment_agent(db)
                    enrichment.on_classification_complete(results)
                    enrichment.on_email_processed({
                        "sender": to_email,
                        "subject": subject,
                        "extracted_items": invoice_data.get('items', [])
                    })
                    print("  Enrichment: learned from classification")
                except Exception as e:
                    print(f"  Enrichment learning error: {e}")

            # Session 27: Print cost summary
            if COST_TRACKING_ENABLED:
                summary = _cost_tracker.summary()
                if summary:
                    print(f"  {summary}")

            return True
        return False
    except Exception as e:
        print(f"  âŒ Error: {e}")
        return False
