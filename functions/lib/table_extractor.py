"""
Specialized table extraction with cross-verification.
Critical for: invoices, packing lists, tariff tables, Excel files.

When two methods disagree on table structure (different row/column counts),
the result is flagged for review rather than silently returning garbage.

Session 28C — Assignment 20.
NEW FILE — does not modify any existing code.
"""

import io
import logging

logger = logging.getLogger("rcb.table_extractor")


class TableExtractor:
    """
    Uses multiple methods to extract tables, then cross-verifies.
    If methods disagree on structure (columns, rows), flags for review.
    """

    def extract_tables(self, file_bytes, content_type):
        """
        Extract tables from a document using multiple methods.

        Returns:
            dict with keys: tables, confidence, methods_tried, warnings
        """
        results = []
        ct = (content_type or "").lower()

        if "pdf" in ct:
            t = self._pdfplumber_tables(file_bytes)
            if t is not None:
                results.append(("pdfplumber", t))

        elif "spreadsheet" in ct or "excel" in ct:
            t = self._openpyxl_tables(file_bytes)
            if t is not None:
                results.append(("openpyxl", t))

        elif "html" in ct:
            t = self._bs4_tables(file_bytes)
            if t is not None:
                results.append(("beautifulsoup", t))
            t2 = self._pandas_html_tables(file_bytes)
            if t2 is not None:
                results.append(("pandas_html", t2))

        if not results:
            return {
                "tables": [],
                "confidence": 0,
                "methods_tried": [],
                "warnings": ["No tables found"],
            }

        # Cross-verify when we have 2+ results
        warnings = []
        if len(results) >= 2:
            if self._structures_disagree(results[0][1], results[1][1]):
                warnings.append(
                    f"Table methods disagree: '{results[0][0]}' found "
                    f"{self._table_shape(results[0][1])} but "
                    f"'{results[1][0]}' found "
                    f"{self._table_shape(results[1][1])}. "
                    f"Using '{results[0][0]}'. Review recommended."
                )

        return {
            "tables": results[0][1],
            "confidence": 0.9 if not warnings else 0.6,
            "methods_tried": [r[0] for r in results],
            "warnings": warnings,
        }

    # ─────────────────────────────────────
    #  Extraction methods
    # ─────────────────────────────────────

    def _pdfplumber_tables(self, file_bytes):
        """Extract tables from PDF using pdfplumber."""
        try:
            import pdfplumber
            tables = []
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages[:50]:
                    page_tables = page.extract_tables()
                    for tbl in (page_tables or []):
                        parsed = self._to_dicts(tbl)
                        if parsed:
                            tables.append(parsed)
            return tables if tables else None
        except Exception as e:
            logger.debug(f"pdfplumber tables failed: {e}")
            return None

    def _openpyxl_tables(self, file_bytes):
        """Extract tables from Excel using openpyxl."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes), read_only=True, data_only=True,
            )
            tables = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_text = [str(c) if c is not None else "" for c in row]
                    if any(cell.strip() for cell in row_text):
                        rows.append(row_text)
                if rows:
                    parsed = self._to_dicts(rows)
                    if parsed:
                        tables.append(parsed)
            wb.close()
            return tables if tables else None
        except Exception as e:
            logger.debug(f"openpyxl tables failed: {e}")
            return None

    def _bs4_tables(self, file_bytes):
        """Extract tables from HTML using BeautifulSoup."""
        try:
            from bs4 import BeautifulSoup
            encoding = self._detect_encoding(file_bytes)
            html_text = file_bytes.decode(encoding, errors="replace")
            soup = BeautifulSoup(html_text, "html.parser")

            tables = []
            for table_el in soup.find_all("table"):
                rows = []
                for tr in table_el.find_all("tr"):
                    cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
                    if cells:
                        rows.append(cells)
                if rows:
                    parsed = self._to_dicts(rows)
                    if parsed:
                        tables.append(parsed)
            return tables if tables else None
        except Exception as e:
            logger.debug(f"BS4 tables failed: {e}")
            return None

    def _pandas_html_tables(self, file_bytes):
        """Extract tables from HTML using pandas (if available)."""
        try:
            import pandas as pd
            encoding = self._detect_encoding(file_bytes)
            html_text = file_bytes.decode(encoding, errors="replace")
            dfs = pd.read_html(io.StringIO(html_text))
            tables = []
            for df in dfs:
                df = df.fillna("")
                records = df.astype(str).to_dict("records")
                if records:
                    tables.append(records)
            return tables if tables else None
        except ImportError:
            return None  # pandas not installed
        except Exception as e:
            logger.debug(f"pandas HTML tables failed: {e}")
            return None

    # ─────────────────────────────────────
    #  Cross-verification
    # ─────────────────────────────────────

    def _structures_disagree(self, tables1, tables2):
        """Check if two sets of tables have meaningfully different structure."""
        if len(tables1) != len(tables2):
            return True
        for t1, t2 in zip(tables1, tables2):
            if abs(len(t1) - len(t2)) > 2:
                return True
            if t1 and t2:
                if set(t1[0].keys()) != set(t2[0].keys()):
                    return True
        return False

    def _table_shape(self, tables):
        """Human-readable description of table dimensions."""
        shapes = [
            f"{len(t)} rows x {len(t[0]) if t else 0} cols"
            for t in tables
        ]
        return f"{len(tables)} table(s) ({', '.join(shapes)})"

    # ─────────────────────────────────────
    #  Helpers
    # ─────────────────────────────────────

    def _to_dicts(self, rows):
        """Convert raw table rows (list of lists) to list of dicts."""
        if not rows or len(rows) < 2:
            return []
        headers = [
            str(h).strip() if h else f"col_{i}"
            for i, h in enumerate(rows[0])
        ]
        return [
            {
                headers[j] if j < len(headers) else f"col_{j}":
                    str(cell).strip() if cell else ""
                for j, cell in enumerate(row)
            }
            for row in rows[1:]
            if any(cell for cell in row)
        ]

    def _detect_encoding(self, file_bytes):
        """Detect encoding, with Hebrew-friendly fallbacks."""
        for enc in ("utf-8", "windows-1255", "iso-8859-8", "latin-1"):
            try:
                file_bytes[:2000].decode(enc)
                return enc
            except (UnicodeDecodeError, LookupError):
                continue
        return "utf-8"
