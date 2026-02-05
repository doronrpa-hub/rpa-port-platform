import re

# Read current file
with open('main.py', 'r') as f:
    content = f.read()

# Multi-agent code to insert
multiagent_code = '''
# ============================================================
# MULTI-AGENT CLASSIFICATION SYSTEM  
# ============================================================

def call_classification_agent(agent_name, prompt, context, api_key):
    """Call a specialized classification agent"""
    try:
        response = requests.post(
            'https://api.anthropic.com/v1/messages',
            headers={
                'x-api-key': api_key,
                'anthropic-version': '2023-06-01',
                'content-type': 'application/json'
            },
            json={
                'model': 'claude-sonnet-4-20250514',
                'max_tokens': 3000,
                'system': f"You are {agent_name} for R.P.A. PORT LTD Israel customs broker. Respond in Hebrew.",
                'messages': [{'role': 'user', 'content': prompt + chr(10) + context[:12000]}]
            },
            timeout=90
        )
        if response.status_code == 200:
            return response.json().get('content', [{}])[0].get('text', '')
        print(f"{agent_name} error: {response.status_code}")
        return None
    except Exception as e:
        print(f"{agent_name} exception: {e}")
        return None

def run_multi_agent_classification(docs_with_text, subject, body):
    """Run multi-agent classification pipeline"""
    api_key = get_anthropic_key()
    if not api_key:
        return None, None, None
    
    # Build docs text
    docs_text = ""
    for d in docs_with_text:
        if d.get('extracted_text'):
            docs_text += f"\\n=== {d.get('name','')} ===\\n{d['extracted_text'][:10000]}"
    
    if not docs_text:
        return None, None, None
    
    ctx = f"Subject: {subject}\\nDocuments:{docs_text}"
    
    print("Running multi-agent classification...")
    
    # Document Agent
    print("  Document Agent...")
    doc_out = call_classification_agent("Document Agent",
        "Extract: seller, buyer, invoice#, date, incoterms, items (description, qty, price, HS code), total, currency, container#, origin declaration",
        ctx, api_key)
    
    # HS Agent  
    print("  HS Classification Agent...")
    hs_out = call_classification_agent("HS Classification Agent",
        "Classify each product: 1) Identify product/material/use 2) Find Chapter 3) Find Heading (4-digit) using GIR 4) Find subheading (10-digit) 5) Compare to supplier code. Give: HS code, Hebrew description, duty rate, validation",
        ctx, api_key)
    
    # Librarian Agent
    print("  Librarian Agent...")
    lib_out = call_classification_agent("Librarian Agent", 
        "Find relevant: court cases (PIDAN, VIVO, DENVER SANDALS), customs rulings, BTI decisions, classification traps",
        ctx + "\\n\\nHS:" + (hs_out or ""), api_key)
    
    # Regulatory Agent
    print("  Regulatory Agent...")
    reg_out = call_classification_agent("Regulatory Agent",
        "Check: Free Import Order (Schedule 1/2), required permits (SII, MoC, Health, Agriculture), standards",
        ctx, api_key)
    
    # FTA Agent
    print("  FTA Agent...")
    fta_out = call_classification_agent("FTA Agent",
        "Check: origin country, applicable FTA, origin declaration validity, EUR.1 needed, duty savings",
        ctx, api_key)
    
    # Risk Agent
    print("  Risk Agent...")
    risk_out = call_classification_agent("Risk Agent",
        "Identify: classification traps, audit risks, valuation issues, document mismatches, red flags",
        ctx + "\\n\\n" + (hs_out or ""), api_key)
    
    # Synthesis Agent
    print("  Synthesis Agent...")
    all_out = f"DOC:{doc_out}\\n\\nHS:{hs_out}\\n\\nLIB:{lib_out}\\n\\nREG:{reg_out}\\n\\nFTA:{fta_out}\\n\\nRISK:{risk_out}"
    final = call_classification_agent("Synthesis Agent",
        "Combine all agent outputs into final report: a)Transaction details b)Classification table c)Regulatory requirements d)FTA benefits e)Research findings f)Risks g)Financial summary h)Recommendations i)Status",
        all_out, api_key)
    
    if final:
        print("Multi-agent classification complete!")
        excel_data = {'final': final, 'doc': doc_out, 'hs': hs_out, 'lib': lib_out, 'reg': reg_out, 'fta': fta_out, 'risk': risk_out}
        return final, all_out, excel_data
    return None, None, None

def create_classification_excel(data):
    """Create multi-sheet Excel report"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        import io
        import base64
        
        wb = Workbook()
        sheets = [("Summary", data.get('final','')), ("HS Classification", data.get('hs','')), 
                  ("Regulatory", data.get('reg','')), ("FTA", data.get('fta','')),
                  ("Research", data.get('lib','')), ("Risk", data.get('risk','')), ("Documents", data.get('doc',''))]
        
        for i, (name, txt) in enumerate(sheets):
            ws = wb.active if i == 0 else wb.create_sheet(name)
            if i == 0: ws.title = name
            ws['A1'] = name
            ws['A1'].font = Font(size=14, bold=True)
            ws['A3'] = (txt or '')[:30000]
            ws.column_dimensions['A'].width = 100
        
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return base64.b64encode(out.read()).decode('utf-8')
    except Exception as e:
        print(f"Excel error: {e}")
        return None

def build_classification_report_email(report, appendix, sender):
    """Build classification report HTML email"""
    name = sender.split('<')[0].strip().split()[0] if sender else "שלום"
    name = to_hebrew_name(name) if name else "שלום"
    
    return f"""
    <div dir="rtl" style="font-family:Arial;font-size:12pt;line-height:1.6">
        <p><strong>שלום {name},</strong></p>
        <p>סיימתי לנתח את המסמכים. מצורף דו"ח סיווג מקיף:</p>
        <div style="background:#fff;border:2px solid #1e3a5f;border-radius:12px;margin:20px 0">
            <div style="background:#1e3a5f;color:white;padding:15px">
                <h2 style="margin:0">🏷️ דו"ח סיווג מכס - RCB AI</h2>
            </div>
            <div style="padding:20px;white-space:pre-wrap">{report}</div>
        </div>
        <details style="background:#f5f5f5;border-radius:8px;margin:20px 0">
            <summary style="padding:15px;cursor:pointer;font-weight:bold">📚 נספח: פלטי סוכנים</summary>
            <div style="padding:15px;white-space:pre-wrap;font-size:10pt">{appendix or ''}</div>
        </details>
        <p>📎 מצורפים: המסמכים המקוריים + קובץ Excel</p>
        <div style="background:#fff3cd;border:1px solid #ffc107;padding:12px;border-radius:8px;margin:15px 0">
            ⚠️ דו"ח זה הופק ע"י AI ומהווה המלצה בלבד.
        </div>
        <hr style="margin:25px 0">
        <table dir="rtl"><tr>
            <td style="padding-left:15px"><img src="https://rpa-port.com/wp-content/uploads/2020/01/logo.png" style="width:80px"></td>
            <td style="border-right:3px solid #1e3a5f;padding-right:15px">
                <strong style="color:#1e3a5f">🤖 RCB - AI Customs Broker</strong><br>
                <strong>R.P.A. PORT LTD</strong><br>
                <span style="font-size:10pt">📧 rcb@rpa-port.co.il</span>
            </td>
        </tr></table>
    </div>
    """

'''

# Find where to insert - before build_rcb_reply
if 'run_multi_agent_classification' not in content:
    marker = 'def build_rcb_reply'
    if marker in content:
        content = content.replace(marker, multiagent_code + '\n' + marker)
        print("✅ Multi-agent code added!")
    else:
        print("❌ Could not find insertion point")
else:
    print("⚠️ Multi-agent code already exists")

# Write back
with open('main.py', 'w') as f:
    f.write(content)

print(f"File now has {len(content)} characters")
