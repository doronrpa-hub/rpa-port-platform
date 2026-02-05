with open('main.py', 'r') as f:
    content = f.read()

multi_agent_code = '''

# ============================================================
# MULTI-AGENT CLASSIFICATION SYSTEM
# ============================================================

import base64
import io

def extract_text_from_pdf_base64(content_bytes_b64):
    """Extract text from base64-encoded PDF"""
    try:
        pdf_bytes = base64.b64decode(content_bytes_b64)
        import pdfplumber
        text = ""
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages[:20]:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\\n"
        return text[:40000]
    except Exception as e:
        print(f"PDF extraction error: {e}")
        return ""

def call_agent(agent_name, agent_prompt, context, api_key):
    """Call a specialized agent"""
    try:
        response = requests.post(
            'https://api.anthropic.com/v1/messages',
            headers={
                'x-api-key': api_key,
                'anthropic-version': '2023-06-01',
                'content-type': 'application/json'
            },
            json={
                'model': 'claude-sonnet-4-20250514',
                'max_tokens': 3000,
                'system': f"You are the {agent_name} for R.P.A. PORT LTD Israel customs broker. Respond in Hebrew. Be thorough.",
                'messages': [{'role': 'user', 'content': agent_prompt + "\\n\\n" + context[:15000]}]
            },
            timeout=90
        )
        if response.status_code == 200:
            return response.json().get('content', [{}])[0].get('text', '')
        return None
    except Exception as e:
        print(f"{agent_name} error: {e}")
        return None

def run_multi_agent_classification(attachments_with_text, email_subject, email_body):
    """Run all agents and synthesize results"""
    api_key = get_anthropic_key()
    if not api_key:
        return None, None, None
    
    docs_text = ""
    for att in attachments_with_text:
        if att.get('extracted_text'):
            docs_text += f"\\n=== {att.get('name','')} ===\\n{att['extracted_text'][:12000]}"
    
    if not docs_text:
        return None, None, None
    
    context = f"Subject: {email_subject}\\n\\nDocuments:{docs_text}"
    
    print("🎯 Running multi-agent classification...")
    
    # Agent 1: Document Extraction
    print("  📄 Document Agent...")
    doc_prompt = """חלץ מהמסמכים:
- פרטי ספק (שם, כתובת, מדינה)
- פרטי קונה
- מספר חשבון ותאריך
- Incoterms
- רשימת פריטים: תיאור, כמות, מחיר, קוד HS של הספק
- סה"כ ומטבע
- מספר מכולה/B/L
- הצהרת מקור (אם יש)"""
    doc_out = call_agent("Document Agent", doc_prompt, context, api_key)
    
    # Agent 2: HS Classification
    print("  🔢 HS Classification Agent...")
    hs_prompt = """סווג כל מוצר לפי המתודולוגיה:
1. זהה: מה המוצר, חומר, שימוש, טכנולוגיה
2. מצא פרק (Chapter) - בדוק הערות
3. מצא פרט (Heading 4 ספרות) - יישם GIR 1-6
4. מצא תת-פרט (6-10 ספרות)
5. השווה לקוד הספק - אם שונה, הסבר למה

לכל פריט תן:
- קוד HS מומלץ (10 ספרות)
- תיאור עברי
- שיעור מכס
- האם קוד הספק נכון"""
    hs_out = call_agent("HS Classification Agent", hs_prompt, context, api_key)
    
    # Agent 3: Librarian
    print("  📚 Librarian Agent...")
    lib_prompt = """חפש תקדימים רלוונטיים:
- פסקי דין: PIDAN, VIVO, DENVER SANDALS, SAMI COHEN, YAKBI
- החלטות סיווג של רשות המכס
- BTI אירופאי/אמריקאי
- מלכודות סיווג ידועות בתחום"""
    lib_out = call_agent("Librarian Agent", lib_prompt, context + "\\n\\nClassification:\\n" + (hs_out or ""), api_key)
    
    # Agent 4: Regulatory
    print("  ⚖️ Regulatory Agent...")
    reg_prompt = """בדוק דרישות רגולטוריות:
- צו יבוא חופשי: תוספת 1 (חופשי) או תוספת 2 (טעון אישור)
- אישורים נדרשים: SII, MoC, משרד הבריאות, חקלאות
- תקנים ישראליים (ת"י)
- רישיונות יבוא"""
    reg_out = call_agent("Regulatory Agent", reg_prompt, context, api_key)
    
    # Agent 5: FTA & Origin
    print("  🌍 FTA Agent...")
    fta_prompt = """בדוק הסכמי סחר ומקור:
- זהה ארץ מקור
- הסכם FTA רלוונטי (EU, EFTA, US, UK, Turkey, וכו')
- הצהרת מקור על החשבון - האם קיימת ותקינה?
- EUR.1 נדרש?
- הפרש מכס רגיל vs FTA"""
    fta_out = call_agent("FTA Origin Agent", fta_prompt, context, api_key)
    
    # Agent 6: Risk
    print("  🚨 Risk Agent...")
    risk_prompt = """זהה סיכונים:
- מלכודות סיווג
- סיכון ביקורת מכס
- ערך חשוד (נמוך/גבוה מדי)
- אי-התאמות בין מסמכים
- דגלים אדומים"""
    risk_out = call_agent("Risk Agent", risk_prompt, context + "\\n\\n" + (hs_out or ""), api_key)
    
    # Agent 7: Synthesis
    print("  🧠 Synthesis Agent...")
    synth_prompt = """שלב את כל הממצאים לדו"ח סופי:

📊 דו"ח סיווג מכס - RCB AI

א. פרטי עסקה (מסוכן המסמכים)
ב. סיווג מוצרים - טבלה:
   | פריט | תיאור | קוד ספק | קוד מומלץ | מכס | רישיון |
ג. דרישות רגולטוריות
ד. הסכמי סחר והטבות
ה. ממצאי מחקר ותקדימים
ו. התראות וסיכונים
ז. סיכום כספי (מכס + מע"מ משוער)
ח. המלצות
ט. סטטוס: מוכן לשחרור / דורש השלמות"""
    
    all_outputs = f"""
=== DOCUMENT AGENT ===
{doc_out or 'N/A'}

=== HS CLASSIFICATION ===
{hs_out or 'N/A'}

=== LIBRARIAN ===
{lib_out or 'N/A'}

=== REGULATORY ===
{reg_out or 'N/A'}

=== FTA ===
{fta_out or 'N/A'}

=== RISK ===
{risk_out or 'N/A'}
"""
    final_report = call_agent("Synthesis Agent", synth_prompt, all_outputs, api_key)
    
    if final_report:
        print("✅ Multi-agent classification complete!")
        excel_data = {
            'final_report': final_report,
            'doc_output': doc_out,
            'hs_output': hs_out,
            'lib_output': lib_out,
            'reg_output': reg_out,
            'fta_output': fta_out,
            'risk_output': risk_out
        }
        appendix = f"📚 פלטי סוכנים:\\n\\n{all_outputs}"
        return final_report, appendix, excel_data
    
    return None, None, None

def create_multi_sheet_excel(excel_data):
    """Create Excel with sheets per agent"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        wb = Workbook()
        
        sheets = [
            ("סיכום", excel_data.get('final_report', '')),
            ("סיווג HS", excel_data.get('hs_output', '')),
            ("רגולציה", excel_data.get('reg_output', '')),
            ("הסכמי סחר", excel_data.get('fta_output', '')),
            ("מחקר", excel_data.get('lib_output', '')),
            ("סיכונים", excel_data.get('risk_output', '')),
            ("מסמכים", excel_data.get('doc_output', ''))
        ]
        
        for i, (name, content) in enumerate(sheets):
            if i == 0:
                ws = wb.active
                ws.title = name
            else:
                ws = wb.create_sheet(name)
            ws['A1'] = name
            ws['A1'].font = Font(size=14, bold=True)
            ws['A3'] = (content or '')[:30000]
            ws.column_dimensions['A'].width = 100
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return base64.b64encode(output.read()).decode('utf-8')
    except Exception as e:
        print(f"Excel error: {e}")
        return None

def build_classification_email(main_report, appendix, sender_name):
    """Build classification report email"""
    name = sender_name.split('<')[0].strip().split()[0] if sender_name else "שלום"
    name = to_hebrew_name(name) if name else "שלום"
    
    html = f"""
    <div dir="rtl" style="font-family: Arial, sans-serif; font-size: 12pt; line-height: 1.6;">
        <p><strong>שלום {name},</strong></p>
        <p>סיימתי לנתח את המסמכים. מצורף דו"ח סיווג מקיף:</p>
        
        <div style="background: #fff; border: 2px solid #1e3a5f; border-radius: 12px; margin: 20px 0; overflow: hidden;">
            <div style="background: #1e3a5f; color: white; padding: 15px;">
                <h2 style="margin: 0;">🏷️ דו"ח סיווג מכס - RCB AI</h2>
            </div>
            <div style="padding: 20px; white-space: pre-wrap;">{main_report}</div>
        </div>
        
        <details style="background: #f5f5f5; border-radius: 8px; margin: 20px 0;">
            <summary style="padding: 15px; cursor: pointer; font-weight: bold;">📚 נספח: פירוט סוכנים</summary>
            <div style="padding: 15px; white-space: pre-wrap; font-size: 10pt;">{appendix or ''}</div>
        </details>
        
        <p>📎 מצורפים: המסמכים המקוריים + קובץ Excel עם פירוט מלא</p>
        
        <div style="background: #fff3cd; border: 1px solid #ffc107; padding: 12px; border-radius: 8px; margin: 15px 0;">
            ⚠️ דו"ח זה הופק ע"י AI ומהווה המלצה בלבד. הסיווג הסופי כפוף לאישור רשות המכס.
        </div>
        
        <hr style="margin: 25px 0;">
        <table dir="rtl">
            <tr>
                <td style="padding-left: 15px;">
                    <img src="https://rpa-port.com/wp-content/uploads/2020/01/logo.png" style="width: 80px;">
                </td>
                <td style="border-right: 3px solid #1e3a5f; padding-right: 15px;">
                    <strong style="color: #1e3a5f;">🤖 RCB - AI Customs Broker</strong><br>
                    <strong>R.P.A. PORT LTD</strong><br>
                    <span style="font-size: 10pt;">📧 rcb@rpa-port.co.il</span>
                </td>
            </tr>
        </table>
    </div>
    """
    return html

'''

marker = "def build_rcb_reply"
if 'run_multi_agent_classification' not in content and marker in content:
    content = content.replace(marker, multi_agent_code + "\\n\\n" + marker)

with open('main.py', 'w') as f:
    f.write(content)

print("✅ Multi-agent system added!")
